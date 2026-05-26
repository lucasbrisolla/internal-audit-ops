"""
generate_tracker_achados_v4.py
Tracker de Achados v4 — adiciona aba "Planos de Ação (Flat)" que lê e normaliza
uma planilha externa de controle dos planos de ação.

A planilha externa tem estrutura vertical irregular:
  - 1 item = várias linhas (campos do item na linha principal +
    sub-linhas de recomendações externas intercaladas com sub-linhas de PAs)
  - Cada bloco de PA repete cabeçalhos: Plano de Ação N | O que Fazer |
    Entregáveis | Status | Critérios para o Aceite | Observação

Esta aba normaliza para: 1 linha por PA, com campos do item repetidos.

Abas:
  1. Tracker              — visão executiva (11 colunas)
  2. Detalhe 5Cs          — condição, critério, causa, consequência, ação corretiva
  3. Planos de Ação (Flat)— tabela normalizada lida da planilha externa
  4. _Ref_Severidade      — referência de severidade

Uso:
    python3 scripts/generate_tracker_achados_v4.py [saida.xlsx] [fonte_planos.xlsx]
    Defaults:
      saida : templates/planilhas/Tracker_Achados_v4_template.xlsx
      fonte_planos : ~/Downloads/Controle dos Planos de Ação.xlsx
"""
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("openpyxl não instalado. Execute: pip install openpyxl")

# ── Paleta ───────────────────────────────────────────────────────────────
AZUL_HEADER  = "1F4E79"
AZUL_SUB     = "2F75B6"
VERDE_HEADER = "375623"
ROXO_HEADER  = "7030A0"
CINZA_TITULO = "D6DCE4"
BRANCO       = "FFFFFF"
CINZA_ALT    = "F2F2F2"
AMARELO_HINT = "FFF2CC"

# Cores exatas da escala referencial metodológico/IIA (severity.md + iia-ippf.md)
SEVERITY_COLORS = {
    "Crítico (Material Weakness)":           ("FF0000", BRANCO),
    "Alto (Significant Deficiency)":         ("FF9900", "000000"),
    "Moderado (Deficiency)":                 ("FFFF00", "000000"),
    "Baixo (Exception / Obs. de Melhoria)":  ("92D050", "000000"),
}

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=BRANCO, size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")

def _border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _add_dv(ws, col_letter, r_start, r_end, choices):
    formula = '"' + ",".join(choices) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.sqref = f"{col_letter}{r_start}:{col_letter}{r_end}"
    ws.add_data_validation(dv)

def _header_cell(ws, row, col, value, fill_hex, font_color=BRANCO, bold=True, wrap=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _fill(fill_hex)
    cell.font = Font(bold=bold, color=font_color, size=10, name="Arial")
    cell.alignment = _center(wrap=wrap)
    cell.border = _border_thin()
    return cell


# ── Domínio — derivado das fontes lidas ──────────────────────────────────

# Estados: states.yml
ESTADOS_ACHADO = [
    "rascunho", "revisão_técnica", "aprovado",
    "comunicado", "resposta_recebida", "fechado",
]

# Escala IIA (iia-ippf.md §7) + mapeamento referencial metodológico (severity.md §Classificações)
SEVERIDADES = [
    "Crítico (Material Weakness)",
    "Alto (Significant Deficiency)",
    "Moderado (Deficiency)",
    "Baixo (Exception / Obs. de Melhoria)",
]


PROCESS_CODES = [
    "FEC", "FAT", "REC", "CAD-CLI", "ENT", "DEV",
    "CAD-AT", "AQU", "PAT", "ALI", "INV", "TKL",
]

SIM_NAO  = ["Sim", "Não"]
N_ROWS   = 100

TIPO_ACAO = ["Pontual", "Estrutural", "Roadmap"]


# ════════════════════════════════════════════════════════════════════════════
# ABA 1 — Tracker de Achados (visão executiva simplificada)
# 11 colunas: ID | Processo | Título | Severidade | Falha Desenho | Falha Efetiv.
#             | Compensatório | Owner | Prazo | Status | Nota Agregação/Obs
# ════════════════════════════════════════════════════════════════════════════
def build_tracker(wb):
    ws = wb.create_sheet("Tracker de Achados")
    ws.sheet_view.showGridLines = False

    N_COLS = 11

    # título
    ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
    t = ws["A1"]
    t.value = "TRACKER DE ACHADOS — Cliente Exemplo (IIA IPPF 2024)"
    t.fill  = _fill(AZUL_HEADER)
    t.font  = _font(bold=True, size=12)
    t.alignment = _center()
    t.border = _border_thin()
    ws.row_dimensions[1].height = 28

    # grupos de colunas (linha 2)
    grupos = [
        (1,  3,  "IDENTIFICAÇÃO",  AZUL_HEADER),
        (4,  7,  "AVALIAÇÃO",      "C00000"),
        (8,  9,  "REMEDIAÇÃO",     "7030A0"),
        (10, 11, "STATUS",         "833C11"),
    ]
    for c_ini, c_fim, label, fhex in grupos:
        ws.merge_cells(f"{get_column_letter(c_ini)}2:{get_column_letter(c_fim)}2")
        cell = ws.cell(row=2, column=c_ini, value=label)
        cell.fill = _fill(fhex)
        cell.font = _font(bold=True, size=9)
        cell.alignment = _center()
        cell.border = _border_thin()
        for c in range(c_ini + 1, c_fim + 1):
            ws.cell(row=2, column=c).fill = _fill(fhex)
            ws.cell(row=2, column=c).border = _border_thin()
    ws.row_dimensions[2].height = 18

    # cabeçalhos (linha 3)
    headers = [
        # IDENTIFICAÇÃO
        (1,  "ID Achado\n(ex.: FEC-01)",               14, AZUL_HEADER),
        (2,  "Processo",                                14, AZUL_HEADER),
        (3,  "Título do Achado",                        45, AZUL_HEADER),
        # AVALIAÇÃO
        (4,  "Severidade",                              22, "C00000"),
        (5,  "Falha de\nDesenho?",                      14, "C00000"),
        (6,  "Falha de\nEfetividade?",                  14, "C00000"),
        (7,  "Controle\nCompensatório?",                16, "C00000"),
        # REMEDIAÇÃO
        (8,  "Owner\n(Responsável)",                    25, "7030A0"),
        (9,  "Prazo-Alvo",                              14, "7030A0"),
        # STATUS
        (10, "Status",                                  18, "833C11"),
        (11, "Nota de Agregação / Observações",         45, "833C11"),
    ]

    for col, label, width, fhex in headers:
        cell = ws.cell(row=3, column=col, value=label)
        cell.fill = _fill(fhex)
        cell.font = _font(bold=True, size=9)
        cell.alignment = _center(wrap=True)
        cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 40
    ws.freeze_panes = "A4"

    # linhas de dados
    end = 4 + N_ROWS - 1
    for row in range(4, end + 1):
        fill = _fill(CINZA_ALT) if row % 2 == 0 else _fill(BRANCO)
        for col in range(1, N_COLS + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = _border_thin()
            cell.alignment = _left()
            cell.font = Font(size=10, name="Arial", color="000000")

    # validações de lista
    _add_dv(ws, "B", 4, end, PROCESS_CODES)
    _add_dv(ws, "D", 4, end, SEVERIDADES)
    _add_dv(ws, "E", 4, end, SIM_NAO)
    _add_dv(ws, "F", 4, end, SIM_NAO)
    _add_dv(ws, "G", 4, end, SIM_NAO)
    _add_dv(ws, "J", 4, end, ESTADOS_ACHADO)

    # linha de exemplo (baseada em FEC-01)
    ex = [
        "FEC-01", "FEC",
        "Ausência de validação dos preços publicados no site (PF)",
        "Alto (Significant Deficiency)",
        "Sim", "Sim", "Não",
        "Responsável do Processo", "2026-06-20",
        "aprovado",
        "Agrega com FEC-02 e FEC-04 — mesmo componente COSO P10. Ver Detalhe 5Cs.",
    ]
    for col, val in enumerate(ex, 1):
        cell = ws.cell(row=4, column=col, value=val)
        cell.font = Font(size=10, name="Arial", color="595959", italic=True)
        if col == 4:
            fhex, fc = SEVERITY_COLORS.get(val, (BRANCO, "000000"))
            cell.fill = _fill(fhex)
            cell.font = Font(size=10, name="Arial", bold=True, color=fc, italic=False)
        cell.alignment = _left(wrap=True)
    ws.row_dimensions[4].height = 60

    # nota de rodapé
    r_nota = end + 2
    ws.merge_cells(f"A{r_nota}:{get_column_letter(N_COLS)}{r_nota}")
    nota = ws.cell(
        row=r_nota, column=1,
        value=(
            "AGREGAÇÃO: achados no mesmo processo com Falha de Desenho = Sim + Compensatório = Não "
            "devem ser avaliados coletivamente antes de fechar o engagement. "
            "Registrar conclusão na coluna 'Nota de Agregação'. "
            "Detalhe dos 5Cs na aba 'Detalhe 5Cs'."
        ),
    )
    nota.font = Font(size=9, name="Arial", color="595959", italic=True)
    nota.alignment = _left(wrap=True)
    ws.row_dimensions[r_nota].height = 32

    return ws


# ════════════════════════════════════════════════════════════════════════════
# ABA 2 — Detalhe 5Cs (iia-ippf.md §6 + DoD §4)
# Tabela horizontal: uma linha por achado, colunas para cada C.
# ════════════════════════════════════════════════════════════════════════════
def build_detalhe_5c(wb):
    ws = wb.create_sheet("Detalhe 5Cs")
    ws.sheet_view.showGridLines = False

    # ── Título ──────────────────────────────────────────────────────────────
    N_COLS = 10
    last = get_column_letter(N_COLS)
    ws.merge_cells(f"A1:{last}1")
    t = ws["A1"]
    t.value = "DETALHE DOS 5Cs POR ACHADO — IIA GIAS 2024 §15 + DoD §4"
    t.fill  = _fill(AZUL_HEADER)
    t.font  = _font(bold=True, size=12)
    t.alignment = _center()
    t.border = _border_thin()
    ws.row_dimensions[1].height = 28

    # ── Linha de hint (DoD resumido por coluna) ──────────────────────────
    hints = [
        "ID do Tracker\n(ex.: FEC-01)",
        "Processo\n(ex.: FEC)",
        "CONDITION\nFato observado — o que existe hoje.\nNão opinião.",
        "CRITERIA\nNorma/política verificável que define\no que deveria ocorrer. Citar fonte.",
        "CAUSE\nCausa raiz, não sintoma.\nCategoria: ausência de controle / design inadequado / falha humana / falha sistêmica.",
        "CONSEQUENCE\nImpacto real ou potencial.\nDimensão: Financeira / Operacional / Compliance / Fraude.",
        "CORRECTIVE ACTION\nPA numerado: Ação | Owner | Prazo | Tipo (Pontual/Estrutural/Roadmap) | Critério de conclusão.",
        "SEVERITY RATIONALE\nImpacto × probabilidade + compensatório testado + Prudent Official Test.",
        "Status\n(states.yml)",
    ]
    for col, hint in enumerate(hints, 1):
        cell = ws.cell(row=2, column=col, value=hint)
        cell.fill = _fill(AMARELO_HINT)
        cell.font = Font(size=8, name="Arial", color="595959", italic=True)
        cell.alignment = _center(wrap=True)
        cell.border = _border_thin()
    ws.row_dimensions[2].height = 44

    # ── Cabeçalhos (linha 3) ─────────────────────────────────────────────
    N_COLS = 9
    headers = [
        # (label, fill_hex, width)
        ("ID Achado",                                    AZUL_HEADER,  14),
        ("Processo",                                     AZUL_HEADER,  12),
        ("CONDITION\n(Condição)",                        "4472C4",      55),
        ("CRITERIA\n(Critério)",                         "375623",      55),
        ("CAUSE\n(Causa Raiz)",                          "C00000",      55),
        ("CONSEQUENCE\n(Consequência)",                  "7030A0",      55),
        ("CORRECTIVE ACTION\n(Ação Corretiva)",          "833C11",      60),
        ("SEVERITY RATIONALE\n(Racional de Severidade)", AZUL_HEADER,  55),
        ("Status",                                       AZUL_SUB,      18),
    ]

    for col, (label, fhex, width) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=label)
        cell.fill = _fill(fhex)
        cell.font = _font(bold=True, size=10)
        cell.alignment = _center(wrap=True)
        cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 40
    ws.freeze_panes = "A4"

    # ── Linhas de dados ──────────────────────────────────────────────────
    end = 3 + N_ROWS
    for row in range(4, end + 1):
        fill = _fill(CINZA_ALT) if row % 2 == 0 else _fill(BRANCO)
        for col in range(1, N_COLS + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = _border_thin()
            cell.alignment = _left(wrap=True)
            cell.font = Font(size=10, name="Arial", color="000000")
        ws.row_dimensions[row].height = 60

    _add_dv(ws, "B", 4, end, PROCESS_CODES)
    _add_dv(ws, "I", 4, end, ESTADOS_ACHADO)

    # ── Linha de exemplo (FEC-01) ─────────────────────────────────────────────
    ex = [
        "FEC-01", "FEC",
        "Não existe etapa formal que confronte a ata aprovada pelo comitê de pricing com os preços efetivamente publicados no site após cada atualização mensal.",
        "Preços PF devem refletir fielmente a ata do comitê de pricing. Publicação somente após aprovação final. Fonte: seção 1.2.1 da narrativa; COSO Princípio 10.",
        "Ausência de etapa de confronto ata vs. site no processo. Sem responsável definido para verificação pós-publicação. Categoria: Ausência de controle + Design inadequado.",
        "Contratos PF podem ser fechados com preços divergentes da ata — a maior (contestação) ou a menor (perda de margem em todo o ciclo). Dimensão: Financeira, Compliance.",
        (
            "PA1 | Implantar confronto formal ata vs. site por ciclo | Responsável do Processo | 2026-06-20 | Estrutural | "
            "Ata assinada + print do site + e-mail de confirmação arquivados.\n"
            "PA2 | Definir regra: publicação somente após ata assinada | Responsável do Processo | 2026-06-20 | Estrutural | "
            "Procedimento publicado + ata anterior à data de publicação no log do site."
        ),
        "Impacto alto (todos os contratos PF do ciclo). Probabilidade alta (ausência estrutural). Compensatório: inexistente. Prudent Official Test: Alto confirmado.",
        "aprovado",
    ]
    for col, val in enumerate(ex, 1):
        cell = ws.cell(row=4, column=col, value=val)
        cell.font = Font(size=10, name="Arial", color="595959", italic=True)
        cell.alignment = _left(wrap=True)

    return ws


# ════════════════════════════════════════════════════════════════════════════
# ABA 3 — Referência de Severidade (todas as fontes integradas)
# ════════════════════════════════════════════════════════════════════════════
def build_ref_severidade(wb):
    ws = wb.create_sheet("_Ref_Severidade")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "ESCALA DE SEVERIDADE — referencial metodológico App. 10A + IIA IPPF 2024 §7"
    t.fill  = _fill(AZUL_HEADER)
    t.font  = _font(bold=True, size=11)
    t.alignment = _center()
    t.border = _border_thin()
    ws.row_dimensions[1].height = 28

    # ── Tabela 1: 4 níveis referencial metodológico × IIA ──────────────────────────────────
    ws.merge_cells("A3:G3")
    h3 = ws["A3"]
    h3.value = "1. MAPEAMENTO DE REFERÊNCIA (4 NÍVEIS) × IIA IPPF (CRÍTICO/ALTO/MODERADO/BAIXO)"
    h3.fill  = _fill(AZUL_SUB)
    h3.font  = _font(bold=True, size=10)
    h3.alignment = _center()
    h3.border = _border_thin()

    cols_t1 = [
        ("Nível IIA",             14),
        ("Nível referencial metodológico",          28),
        ("Definição (referencial metodológico)",    55),
        ("Chart Aplicável",       16),
        ("Comunicação Obrigatória", 28),
        ("Prazo Máximo Indicativo", 20),
        ("Exemplo",               40),
    ]
    for i, (h, w) in enumerate(cols_t1, 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.fill = _fill(AZUL_HEADER)
        cell.font = _font(bold=True, size=9)
        cell.alignment = _center(wrap=True)
        cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[4].height = 32

    # dados: derivados de severity.md §Classificações + iia-ippf.md §7
    tabela1 = [
        (
            "Crítico",
            "Material Weakness (MW)",
            "Possibilidade razoável (reasonably possible) de distorção material não prevenida "
            "ou não detectada tempestivamente pelos controles da entidade.",
            "Chart 2 (Step 2) / Chart 3 / Chart 4",
            "Nível acima + Governança + Board",
            "30 dias",
            "Fraude por alta administração; reapresentação de DF; "
            "distorção material descoberta pelo auditor que os controles não detectariam.",
        ),
        (
            "Alto",
            "Significant Deficiency (SD)",
            "Menos severa que MW, mas importante o suficiente para merecer atenção da "
            "governança. Magnitude potencial > inconsequente, sem mitigação completa.",
            "Chart 2 (Step 1) / Chart 3 / Chart 4",
            "Nível acima + Governança",
            "90 dias",
            "SoD ausente em processo de alto volume com compensatório parcial; "
            "ITGC deficiente que afeta controle de aplicação relevante.",
        ),
        (
            "Moderado",
            "Deficiency",
            "Falha entre Exception e SD. Ampla faixa de combinações likelihood × magnitude "
            "que não atingem o limiar de SD.",
            "Chart 1 / Chart 2",
            "Responsável + Nível acima",
            "180 dias",
            "Controle manual com 1 exceção isolada em amostra grande; "
            "evidência de execução incompleta mas magnitude potencial baixa.",
        ),
        (
            "Baixo",
            "Exception / Observação de Melhoria",
            "Desvio isolado que não chega a ser Deficiency. Pode sinalizar deficiência "
            "quando combinado com outros desvios — monitorar para agregação.",
            "Chart 1 (avaliação inicial)",
            "Responsável",
            "360 dias",
            "1 item em 40 sem assinatura em controle de baixa materialidade; "
            "oportunidade de automação sem impacto atual identificado.",
        ),
    ]

    for row_i, row_data in enumerate(tabela1, start=5):
        nivel_iia = row_data[0]
        fhex, fc = {
            "Crítico": ("FF0000", BRANCO),
            "Alto":    ("FF9900", "000000"),
            "Moderado":("FFFF00", "000000"),
            "Baixo":   ("92D050", "000000"),
        }[nivel_iia]
        for col_i, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            if col_i == 1:
                cell.fill = _fill(fhex)
                cell.font = Font(bold=True, size=10, name="Arial", color=fc)
                cell.alignment = _center()
            elif col_i == 2:
                cell.fill = _fill(fhex)
                cell.font = Font(bold=True, size=10, name="Arial", color=fc)
                cell.alignment = _left(wrap=True)
            else:
                cell.fill = _fill(CINZA_ALT) if row_i % 2 == 0 else _fill(BRANCO)
                cell.font = Font(size=9, name="Arial", color="000000")
                cell.alignment = _left(wrap=True)
            cell.border = _border_thin()
        ws.row_dimensions[row_i].height = 55

    # ── Tabela 2: Upper Limit (severity.md §Magnitude) ───────────────────
    r2 = 11
    ws.merge_cells(f"A{r2}:G{r2}")
    h2 = ws.cell(row=r2, column=1,
                  value="2. UPPER LIMIT DEVIATION RATE — App. 10B (referencial metodológico) — 90% de confiança")
    h2.fill = _fill(AZUL_SUB)
    h2.font = _font(bold=True, size=10)
    h2.alignment = _center()
    h2.border = _border_thin()

    ul_headers = ["N amostral", "0 desvios", "1 desvio", "2 desvios",
                  "Como usar", "", ""]
    ul_widths   = [14, 14, 14, 14, 55, 14, 14]
    for i, (h, w) in enumerate(zip(ul_headers, ul_widths), 1):
        cell = ws.cell(row=r2 + 1, column=i, value=h)
        cell.fill = _fill(AZUL_HEADER)
        cell.font = _font(bold=True, size=9)
        cell.alignment = _center(wrap=True)
        cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(i)].width = w

    # dados da tabela (severity.md §Upper Limit)
    ul_data = [
        (25, "8,8%", "14,7%", "20,0%"),
        (30, "7,4%", "12,4%", "16,8%"),
        (40, "5,6%",  "9,4%", "12,8%"),
        (50, "4,6%",  "7,6%", "10,3%"),
    ]
    instrucao = (
        "Fórmula: Adjusted Exposure = Gross Exposure × Upper Limit Rate. "
        "Comparar com materialidade para classificar MW vs SD. "
        "Ex.: N=30, 1 desvio → 12,4%. Gross Exposure = R$1.000.000 → "
        "Adjusted = R$124.000. Se < materialidade → SD (não MW). "
        "ATENÇÃO: upper limit e controles compensatórios são abordagens alternativas — "
        "não aplicar as duas ao mesmo desvio."
    )
    for idx, (n, d0, d1, d2) in enumerate(ul_data):
        r_ul = r2 + 2 + idx
        fill = _fill(CINZA_ALT) if idx % 2 == 0 else _fill(BRANCO)
        for col_i, val in enumerate([n, d0, d1, d2], 1):
            cell = ws.cell(row=r_ul, column=col_i, value=val)
            cell.fill = fill
            cell.font = Font(size=10, name="Arial", color="000000",
                             bold=(col_i == 1))
            cell.alignment = _center()
            cell.border = _border_thin()
        if idx == 0:
            ws.merge_cells(f"E{r_ul}:G{r_ul + len(ul_data) - 1}")
            instr_cell = ws.cell(row=r_ul, column=5, value=instrucao)
            instr_cell.fill = _fill(AMARELO_HINT)
            instr_cell.font = Font(size=9, name="Arial", color="595959", italic=True)
            instr_cell.alignment = _left(wrap=True)
            instr_cell.border = _border_thin()

    # ── Tabela 3: Prudent Official Test + Red Flags ───────────────────────
    r3 = r2 + 8
    ws.merge_cells(f"A{r3}:G{r3}")
    h3b = ws.cell(row=r3, column=1,
                   value="3. PRUDENT OFFICIAL TEST + RED FLAGS (severity.md §Prudent Official Test)")
    h3b.fill = _fill(ROXO_HEADER)
    h3b.font = _font(bold=True, size=10)
    h3b.alignment = _center()
    h3b.border = _border_thin()

    ws.merge_cells(f"A{r3+1}:G{r3+1}")
    pot = ws.cell(row=r3+1, column=1,
                  value=(
                      "PERGUNTA OBRIGATÓRIA antes de qualquer conclusão de severidade: "
                      "'Um profissional prudente, com o mesmo conhecimento dos fatos e circunstâncias, "
                      "concordaria com minha conclusão?' "
                      "Gate funciona nos dois sentidos (elevar e rebaixar). "
                      "Teste de realidade: 'se lida no Wall Street Journal, faria sentido para um leigo de negócios?'"
                  ))
    pot.font = Font(size=10, name="Arial", color="000000")
    pot.fill = _fill(AMARELO_HINT)
    pot.alignment = _left(wrap=True)
    pot.border = _border_thin()
    ws.row_dimensions[r3+1].height = 44

    red_flags = [
        "Classificar a falha olhando apenas o erro identificado, não o risco potencial ('could').",
        "Aceitar controle compensatório fraco ou não testado como mitigação suficiente.",
        "Concluir que a falha 'não é grave' sem ligação clara com magnitude e probabilidade.",
        "Subestimar severidade quando não há distorção materializada — viés documentado: 70% das SDs foram subestimadas pela gestão (estudo empírico de referência 2004-05).",
        "Citar monitoramento como compensação sem evidência de que funcionou na prática (COSO publicou guia específico em 2009).",
        "Tratar ITGC isoladamente sem avaliar impacto nos controles de aplicação dependentes.",
        "Aceitar 'razões diferentes' para múltiplos desvios sem avaliar o padrão subjacente (35 desvios em 58 itens não é coincidência).",
    ]
    ws.merge_cells(f"A{r3+2}:G{r3+2}")
    rf_hdr = ws.cell(row=r3+2, column=1, value="RED FLAGS — sinais de subestimação de severidade:")
    rf_hdr.fill = _fill(AZUL_HEADER)
    rf_hdr.font = _font(bold=True, size=10)
    rf_hdr.alignment = _left()
    rf_hdr.border = _border_thin()

    for rf_i, rf_text in enumerate(red_flags, start=r3+3):
        ws.merge_cells(f"A{rf_i}:G{rf_i}")
        cell = ws.cell(row=rf_i, column=1, value=f"⚠  {rf_text}")
        cell.fill = _fill(CINZA_ALT) if rf_i % 2 == 0 else _fill(BRANCO)
        cell.font = Font(size=9, name="Arial", color="000000")
        cell.alignment = _left(wrap=True)
        cell.border = _border_thin()
        ws.row_dimensions[rf_i].height = 22

    # ── Situações presumidamente mais severas (severity.md §Deficiências especiais) ──
    r4 = r3 + 3 + len(red_flags) + 1
    ws.merge_cells(f"A{r4}:G{r4}")
    h4 = ws.cell(row=r4, column=1,
                  value="4. SITUAÇÕES PRESUMIDAMENTE SD OU MW (severity.md §Deficiências Especiais)")
    h4.fill = _fill("C00000")
    h4.font = _font(bold=True, size=10)
    h4.alignment = _center()
    h4.border = _border_thin()

    especiais = [
        "Fraude identificada de qualquer magnitude por parte da alta administração.",
        "Reapresentação de demonstrações financeiras por erro material passado.",
        "Distorção material descoberta pelo auditor que os controles da entidade não teriam detectado.",
        "Supervisão ineficaz do ICFR pelo comitê de auditoria.",
        "Ausência de expertise contábil interna para aplicar GAAP em transações da entidade.",
        "Ausência de programas e controles antifraude.",
        "Controles insuficientes sobre transações não rotineiras.",
        "Controles insuficientes sobre o processo de fechamento periódico.",
        "Indiferença gerencial em corrigir SDs ou MWs já conhecidas.",
    ]
    for esp_i, esp_text in enumerate(especiais, start=r4+1):
        ws.merge_cells(f"A{esp_i}:G{esp_i}")
        cell = ws.cell(row=esp_i, column=1, value=f"•  {esp_text}")
        cell.fill = _fill(CINZA_ALT) if esp_i % 2 == 0 else _fill(BRANCO)
        cell.font = Font(size=9, name="Arial", color="000000")
        cell.alignment = _left(wrap=True)
        cell.border = _border_thin()
        ws.row_dimensions[esp_i].height = 18

    return ws


# ════════════════════════════════════════════════════════════════════════════
# PARSER — lê planilha externa e retorna lista flat de planos de ação
# ════════════════════════════════════════════════════════════════════════════
def parse_plano_acao(source_path: Path) -> list[dict]:
    """
    Lê a aba 'Controle de Planos de Ação' de uma planilha externa.
    Retorna lista de dicts — um por plano de ação.

    Estrutura real por item:
      Linha de item    : col B = ID numérico; col C = Processo; cols K-Q = PA1
      Linha header sec : col B = "", col C = "Recomendação Externa",
                         col K = "Cobertura das Recomendações", col L = "Plano de Ação N"
      Linha rec+PA     : col B = "", col C = nº rec (dígito), col D = texto da recomendação externa,
                         col K = cobertura, col L = título PA, cols M-Q = dados PA
      ...
      Linha separadora : col B = "ID" (cabeçalho repetido)
    """
    workbook = openpyxl.load_workbook(source_path, data_only=True)
    ws = workbook["Controle de Planos de Ação"]
    rows = list(ws.iter_rows(values_only=True))

    def v(cell): return str(cell or "").strip()

    PA_HEADER_PREFIXES = ("Plano de Ação", "O que Fazer")

    records      = []
    ctx          = {}   # campos do item corrente
    next_pa_num  = 1    # número do próximo PA esperado
    external_recommendations = []   # lista de "N. texto" do item corrente
    last_item_id = 0    # maior ID de item visto (garante sentido crescente)

    for row in rows:
        col_b = v(row[1])   # B  — ID item numérico | "ID" | ""
        col_c = v(row[2])   # C  — Processo (linha item) | nº recomendação (dígito) | "Recomendação Externa" | "Conclusão"
        col_d = v(row[3])   # D  — texto da recomendação externa
        col_k = v(row[10])  # K  — Cobertura das Recomendações
        col_l = v(row[11])  # L  — Título do PA | "Plano de Ação N" | ""
        col_m = v(row[12])  # M  — O que Fazer
        col_n = v(row[13])  # N  — Entregáveis
        col_o = v(row[14])  # O  — Status
        col_p = v(row[15])  # P  — Critério de Aceite
        col_q = v(row[16])  # Q  — Observação

        # ── linha de novo item: col B é inteiro crescente ──
        if col_b.isdigit() and int(col_b) > last_item_id:
            last_item_id = int(col_b)
            ctx = {
                "id_item":     col_b,
                "processo":    col_c,
                "subprocesso": v(row[3]),
                "responsavel": v(row[4]),
                "fragilidade": v(row[5]),
                "descricao":   v(row[6]),
                "criticidade": v(row[7]),
                "objetivo":    v(row[8]),
            }
            next_pa_num = 1
            external_recommendations = []
            # PA1 inline nessa mesma linha (col K preenchido e não é header)
            if col_l and col_l not in PA_HEADER_PREFIXES and col_k and \
                    col_k not in ("Cobertura das Recomendações",):
                records.append({**ctx,
                    "pa_num":            "1",
                    "cobertura_rec":     col_k,
                    "pa_titulo":         col_l,
                    "o_que_fazer":       col_m,
                    "entregaveis":       col_n,
                    "status":            col_o,
                    "criterio_aceite":   col_p,
                    "observacao":        col_q,
                    "recomendacoes_externas": "",
                })
                next_pa_num = 2
            continue

        # ── sem contexto ainda ──
        if not ctx:
            continue

        # ── capturar recomendação externa se presente nesta linha (col C = dígito, col D = texto) ──
        if col_c.isdigit() and col_d:
            external_recommendations.append(f"{col_c}. {col_d}")
            recs_str = "\n".join(external_recommendations)
            # backfill nos PAs já emitidos deste item
            for rec in reversed(records):
                if rec["id_item"] == ctx["id_item"]:
                    rec["recomendacoes_externas"] = recs_str
                else:
                    break

        # ── emitir PA se col K = cobertura e col L = título (não header) ──
        if col_l and col_l not in PA_HEADER_PREFIXES and col_k and \
                col_k not in ("Cobertura das Recomendações",):
            records.append({**ctx,
                "pa_num":            str(next_pa_num),
                "cobertura_rec":     col_k,
                "pa_titulo":         col_l,
                "o_que_fazer":       col_m,
                "entregaveis":       col_n,
                "status":            col_o,
                "criterio_aceite":   col_p,
                "observacao":        col_q,
                "recomendacoes_externas": "\n".join(external_recommendations),
            })
            next_pa_num += 1

    return records


# ════════════════════════════════════════════════════════════════════════════
# ABA 3 — Planos de Ação (Flat)
# ════════════════════════════════════════════════════════════════════════════
def build_planos_flat(wb, records: list[dict]):
    ws = wb.create_sheet("Planos de Ação (Flat)")
    ws.sheet_view.showGridLines = False

    # grupos e headers
    # Colunas: ID | Processo | Subprocesso | Responsável | Fragilidade | Criticidade
    #          | Cobertura Rec | PA Título | O que Fazer | Entregáveis
    #          | Status | Critério de Aceite | Observação | Recomendações Externas
    N_COLS = 14

    ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
    t = ws["A1"]
    t.value = "PLANOS DE AÇÃO — TABELA NORMALIZADA (fonte: planilha externa de controle)"
    t.fill  = _fill(AZUL_HEADER)
    t.font  = _font(bold=True, size=12)
    t.alignment = _center()
    t.border = _border_thin()
    ws.row_dimensions[1].height = 28

    grupos = [
        (1,  6,  "ITEM",              AZUL_HEADER),
        (7,  9,  "PLANO DE AÇÃO",     "7030A0"),
        (10, 10, "ENTREGÁVEIS",       "375623"),
        (11, 11, "STATUS",            "833C11"),
        (12, 13, "ACEITE / OBS",      "4472C4"),
        (14, 14, "RECOMENDAÇÕES EXTERNAS", "C55A11"),
    ]
    for c_ini, c_fim, label, fhex in grupos:
        ws.merge_cells(f"{get_column_letter(c_ini)}2:{get_column_letter(c_fim)}2")
        cell = ws.cell(row=2, column=c_ini, value=label)
        cell.fill = _fill(fhex)
        cell.font = _font(bold=True, size=9)
        cell.alignment = _center()
        cell.border = _border_thin()
        for c in range(c_ini + 1, c_fim + 1):
            ws.cell(row=2, column=c).fill = _fill(fhex)
            ws.cell(row=2, column=c).border = _border_thin()
    ws.row_dimensions[2].height = 18

    headers = [
        (1,  "ID",                    6,  AZUL_HEADER),
        (2,  "Processo",              14, AZUL_HEADER),
        (3,  "Subprocesso",           22, AZUL_HEADER),
        (4,  "Responsável",           18, AZUL_HEADER),
        (5,  "Fragilidade",           40, AZUL_HEADER),
        (6,  "Criticidade",           12, AZUL_HEADER),
        (7,  "Cobertura\nRec. Ext.",   14, "7030A0"),
        (8,  "Título do PA",          35, "7030A0"),
        (9,  "O que Fazer",           55, "7030A0"),
        (10, "Entregáveis",           45, "375623"),
        (11, "Status",                16, "833C11"),
        (12, "Critério de Aceite",    45, "4472C4"),
        (13, "Observação",            35, "4472C4"),
        (14, "Recomendações Externas",     55, "C55A11"),
    ]
    for col, label, width, fhex in headers:
        cell = ws.cell(row=3, column=col, value=label)
        cell.fill = _fill(fhex)
        cell.font = _font(bold=True, size=9)
        cell.alignment = _center(wrap=True)
        cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 36
    ws.freeze_panes = "A4"

    STATUS_COLORS = {
        "Não Iniciado": ("FFF2CC", "000000"),
        "Em Andamento": ("DDEBF7", "000000"),
        "Concluído":    ("E2EFDA", "000000"),
        "Atrasado":     ("FCE4D6", "000000"),
    }

    CRIT_COLORS = {
        "Alta":   ("FF0000", BRANCO),
        "Média":  ("FF9900", "000000"),
        "Baixa":  ("92D050", "000000"),
    }

    for i, rec in enumerate(records):
        row = i + 4
        fill_base = _fill(CINZA_ALT) if i % 2 == 0 else _fill(BRANCO)

        vals = [
            rec["id_item"], rec["processo"], rec["subprocesso"],
            rec["responsavel"], rec["fragilidade"], rec["criticidade"],
            rec["cobertura_rec"], rec["pa_titulo"], rec["o_que_fazer"],
            rec["entregaveis"], rec["status"],
            rec["criterio_aceite"], rec["observacao"],
            rec.get("recomendacoes_externas", ""),
        ]

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = _border_thin()
            cell.alignment = _left(wrap=True)

            # colorir criticidade
            if col == 6 and val in CRIT_COLORS:
                fhex, fc = CRIT_COLORS[val]
                cell.fill = _fill(fhex)
                cell.font = Font(size=10, name="Arial", bold=True, color=fc)
            # colorir status
            elif col == 11 and val in STATUS_COLORS:
                fhex, fc = STATUS_COLORS[val]
                cell.fill = _fill(fhex)
                cell.font = Font(size=10, name="Arial", color=fc)
            else:
                cell.fill = fill_base
                cell.font = Font(size=10, name="Arial", color="000000")

        ws.row_dimensions[row].height = 55

    # validação de status
    end = len(records) + 4
    _add_dv(ws, "K", 4, max(end, 4 + N_ROWS),
            ["Não Iniciado", "Em Andamento", "Concluído", "Atrasado"])

    return ws


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════
def main():
    script_dir = Path(__file__).parent
    repo_root  = script_dir.parent
    default_out = repo_root / "templates" / "planilhas" / "Tracker_Achados_v4_template.xlsx"
    default_source = Path.home() / "Downloads" / "Controle dos Planos de Ação.xlsx"

    out_path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_out
    source_path = Path(sys.argv[2]) if len(sys.argv) > 2 else default_source

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_tracker(wb)
    build_detalhe_5c(wb)

    if source_path.exists():
        records = parse_plano_acao(source_path)
        build_planos_flat(wb, records)
        print(f"Fonte externa: {len(records)} planos de ação normalizados")
    else:
        print(f"Aviso: fonte externa não encontrada em {source_path} — aba 'Planos de Ação (Flat)' omitida")

    build_ref_severidade(wb)

    wb.save(out_path)
    print(f"Template gerado: {out_path}")


if __name__ == "__main__":
    main()
