"""
generate_papel_teste_v3.py
Papel de Trabalho — Teste de Controle, versão 3 slim.

Alterações em relação à v2:
  - "Referência do Workpaper (WP Ref)" → "ID do Teste" (operacional para IA interna)
  - "Dependência IPE/ITGC" + "Sistema / Relatório IPE" → linha única
    "IPE / Sistema (se IT-dependent ou Application control)"

Abas:
  1. Cabeçalho do Teste   — identificação, atributos, população, amostra, pré-def exceção
  2. Resultados por Item  — registro item a item com atributos e exceções
  3. _Ref_Amostragem      — N=F/T, fatores de confiança, low-freq, two-stage sequential

Uso:
    python3 scripts/generate_papel_teste_v3.py [saida.xlsx]
    Default: templates/planilhas/Papel_Teste_Controle_v3_template.xlsx
"""
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("openpyxl não instalado. Execute: pip install openpyxl")

# ── Paleta canônica ──────────────────────────────────────────────────────────
AZUL_HEADER  = "1F4E79"
AZUL_SUB     = "2F75B6"
VERDE_HEADER = "375623"
ROXO_HEADER  = "7030A0"
CINZA_TITULO = "D6DCE4"
BRANCO       = "FFFFFF"
CINZA_ALT    = "F2F2F2"
AMARELO_HINT = "FFF2CC"

# ── Helpers ──────────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=BRANCO, size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")

def _border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _add_dv(ws, col_letter, r_start, r_end, choices):
    formula = '"' + ",".join(choices) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.sqref = f"{col_letter}{r_start}:{col_letter}{r_end}"
    ws.add_data_validation(dv)

def _section_label(ws, row, text, last_col, fill_hex=AZUL_SUB):
    lc = get_column_letter(last_col)
    ws.merge_cells(f"A{row}:{lc}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(fill_hex)
    c.font = _font(bold=True, size=10)
    c.alignment = _left(wrap=False)
    c.border = _border_thin()
    ws.row_dimensions[row].height = 22
    for col in range(2, last_col + 1):
        ws.cell(row=row, column=col).border = _border_thin()

def _field_row(ws, row, label, col_label, col_value, col_end,
               fill_label="EBF3FB", height=None):
    lc = ws.cell(row=row, column=col_label, value=label)
    lc.fill = _fill(fill_label)
    lc.font = _font(bold=True, color="000000", size=10)
    lc.border = _border_thin()
    lc.alignment = _left(wrap=False)

    if col_end > col_value:
        ws.merge_cells(
            start_row=row, start_column=col_value,
            end_row=row, end_column=col_end
        )
    vc = ws.cell(row=row, column=col_value)
    vc.fill = _fill(BRANCO)
    vc.border = _border_thin()
    vc.alignment = _left()
    vc.font = Font(size=10, name="Arial", color="000000")
    if height:
        ws.row_dimensions[row].height = height
    return vc

def _hint_merged(ws, row, text, last_col):
    lc = get_column_letter(last_col)
    ws.merge_cells(f"A{row}:{lc}{row}")
    h = ws.cell(row=row, column=1, value=text)
    h.fill  = _fill(AMARELO_HINT)
    h.font  = Font(size=9, name="Arial", color="595959", italic=True)
    h.alignment = _left(wrap=True)
    h.border = _border_thin()
    ws.row_dimensions[row].height = 28
    for col in range(2, last_col + 1):
        ws.cell(row=row, column=col).border = _border_thin()


# ── Domínio canônico ─────────────────────────────────────────────────────────
# control-types-and-reliance.md §1
TIPO_CTRL = [
    "Manual",
    "IT-dependent",
    "Application control (Automatizado)",
    "Híbrido",
]
PREV_DET = ["Preventivo", "Detectivo", "Preventivo/Detectivo"]
FREQ_CTRL = [
    "Diário", "Semanal", "Quinzenal", "Mensal",
    "Trimestral", "Semestral", "Anual", "Sob Demanda",
]
TIPO_TESTE = [
    "Teste de Desenho (ToD)",
    "Teste de Efetividade Operacional (ToE)",
    "ToD + ToE",
]
RESULT_FINAL = ["Efetivo", "Inefetivo", "Inconclusivo"]
RESULT_ITEM  = ["Aprovado", "Exceção", "N/A"]
# sample-size.md — causa de exceção alinhada ao método
CAUSA_EXCEP  = [
    "Falha de Desenho — controle não cobria o risco",
    "Falha de Efetividade — controle existia mas não operou",
    "Falha de Frequência — controle não executado no período",
    "Evidência Insuficiente — não comprovável",
    "Erro Isolado — desvio único sem padrão sistêmico",
    "Possível Override de Gestão",
    "Indicativo de Fraude",
    "Outra",
]
SIM_NAO = ["Sim", "Não", "N/A"]
CONFIANCA_OPTS = [
    "90% (F=2,31)",
    "95% (F=3,00)",
    "87% (F=2,00)",
    "80% (F=1,61)",
]
TOLERANCIA_OPTS = ["5%", "10%", "7%", "3%"]

LAST_COL = 8  # colunas A-H no Cabeçalho


# ════════════════════════════════════════════════════════════════════════════
# ABA 1 — Cabeçalho do Teste
# ════════════════════════════════════════════════════════════════════════════
def build_cabecalho(wb):
    ws = wb.create_sheet("Cabeçalho do Teste")
    ws.sheet_view.showGridLines = False

    # larguras
    widths = [24, 28, 20, 28, 24, 28, 20, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # título
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "PAPEL DE TRABALHO — TESTE DE CONTROLE  v3"
    t.fill  = _fill(AZUL_HEADER)
    t.font  = _font(bold=True, size=13)
    t.alignment = _center()
    t.border = _border_thin()
    ws.row_dimensions[1].height = 32

    r = 3

    # ── IDENTIFICAÇÃO ────────────────────────────────────────────────────
    _section_label(ws, r, "IDENTIFICAÇÃO", LAST_COL)
    r += 1
    _field_row(ws, r, "ID do Teste",        1, 2, 4)
    _field_row(ws, r, "Cliente / Entidade", 5, 6, 8)
    r += 1
    _field_row(ws, r, "Engagement / Trabalho",            1, 2, 4)
    _field_row(ws, r, "Período Auditado",                 5, 6, 8)
    r += 1
    _field_row(ws, r, "Processo",                         1, 2, 4)
    _field_row(ws, r, "Subprocesso",                      5, 6, 8)
    r += 1
    _field_row(ws, r, "Auditor Responsável",              1, 2, 4)
    _field_row(ws, r, "Data de Execução",                 5, 6, 8)
    r += 1
    _field_row(ws, r, "Revisor",                          1, 2, 4)
    _field_row(ws, r, "Data de Revisão",                  5, 6, 8)
    r += 2

    # ── ATRIBUTOS DO CONTROLE (SCOT) ──────────────────────────────────────
    _section_label(ws, r, "ATRIBUTOS DO CONTROLE (SCOT)", LAST_COL)
    r += 1
    _hint_merged(ws, r,
        "Tipo canônico: Manual | IT-dependent | Application control (Automatizado) | Híbrido. "
        "Para IT-dependent e Application controls, preencher 'IPE / Sistema' obrigatoriamente.",
        LAST_COL)
    r += 1
    _field_row(ws, r, "Cód. Controle",      1, 2, 4)
    _field_row(ws, r, "Cód. Risco Coberto", 5, 6, 8)
    r += 1
    _field_row(ws, r, "Descrição do Controle", 1, 2, 8, height=55)
    r += 1
    _field_row(ws, r, "Frequência",            1, 2, 4)
    _field_row(ws, r, "Tipo de Controle",      5, 6, 8)
    _add_dv(ws, "B", r, r, FREQ_CTRL)
    _add_dv(ws, "F", r, r, TIPO_CTRL)
    r += 1
    _field_row(ws, r, "Natureza (Prev./Det.)", 1, 2, 4)
    _field_row(ws, r, "Owner (Executor)",      5, 6, 8)
    _add_dv(ws, "B", r, r, PREV_DET)
    r += 1
    _field_row(ws, r, "IPE / Sistema (se IT-dependent ou Application control)", 1, 2, 8)
    r += 2

    # ── OBJETIVO E CRITÉRIO DO TESTE ─────────────────────────────────────
    _section_label(ws, r, "OBJETIVO E CRITÉRIO DO TESTE", LAST_COL)
    r += 1
    _field_row(ws, r, "Tipo de Teste",                       1, 2, 4)
    _field_row(ws, r, "Critério de Aprovação / Reprovação",  5, 6, 8)
    _add_dv(ws, "B", r, r, TIPO_TESTE)
    r += 1
    _field_row(ws, r, "Objetivo do Teste (o que se quer verificar)", 1, 2, 8, height=55)
    r += 2

    # ── POPULAÇÃO E AMOSTRA ──────────────────────────────────────────────
    _section_label(ws, r, "POPULAÇÃO E AMOSTRA", LAST_COL)
    r += 1
    _hint_merged(ws, r,
        "N = F / T  (F = fator de confiança; T = taxa de tolerância como decimal). "
        "Consulte _Ref_Amostragem para a tabela completa de fatores e valores por frequência.",
        LAST_COL)
    r += 1
    _field_row(ws, r, "Descrição da População",   1, 2, 8, height=44)
    r += 1
    _field_row(ws, r, "Período da População",      1, 2, 4)
    _field_row(ws, r, "Total da População (N)",    5, 6, 8)
    r += 1
    _field_row(ws, r, "Fonte / IPE da População",  1, 2, 4)
    _field_row(ws, r, "Tamanho da Amostra (n)",    5, 6, 8)
    r += 1
    _field_row(ws, r, "Confiança de Amostragem",   1, 2, 4)
    _field_row(ws, r, "Tolerância de Desvio (%)",  5, 6, 8)
    _add_dv(ws, "B", r, r, CONFIANCA_OPTS)
    _add_dv(ws, "F", r, r, TOLERANCIA_OPTS)
    r += 1
    _field_row(ws, r, "Lógica de Seleção da Amostra", 1, 2, 8, height=40)
    r += 2

    # ── DEFINIÇÃO DE EXCEÇÃO (pré-execução) — DoD obrigatório ───────────
    _section_label(ws, r, "DEFINIÇÃO DE EXCEÇÃO — preencher ANTES da execução (DoD)",
                   LAST_COL, fill_hex=ROXO_HEADER)
    r += 1
    _hint_merged(ws, r,
        "Defina o que constitui exceção ANTES de iniciar o teste. "
        "Alterar a definição após ver os resultados invalida a conclusão (viés de confirmação).",
        LAST_COL)
    r += 1
    _field_row(ws, r, "Definição de Exceção\n(o que configura desvio neste teste)", 1, 2, 8, height=55)
    r += 1
    _field_row(ws, r, "Atributos a Verificar por Item (lista dos checks)", 1, 2, 8, height=55)
    r += 2

    # ── PROCEDIMENTO EXECUTADO ───────────────────────────────────────────
    _section_label(ws, r, "PROCEDIMENTO EXECUTADO", LAST_COL)
    r += 1
    _field_row(ws, r, "Descrição do Procedimento (nível reproduzível)", 1, 2, 8, height=80)
    r += 1
    _field_row(ws, r, "Evidência Obtida (referência / localização)",    1, 2, 8, height=40)
    r += 2

    # ── CONCLUSÃO ────────────────────────────────────────────────────────
    _section_label(ws, r, "CONCLUSÃO", LAST_COL, fill_hex=VERDE_HEADER)
    r += 1
    _field_row(ws, r, "Resultado Final do Teste",          1, 2, 4)
    _field_row(ws, r, "Nº de Exceções Identificadas",      5, 6, 8)
    _add_dv(ws, "B", r, r, RESULT_FINAL)
    r += 1
    _field_row(ws, r, "Narrativa da Conclusão (resultado × risco coberto)", 1, 2, 8, height=60)
    r += 1
    _field_row(ws, r, "Análise Qualitativa das Exceções (isolado / sistêmico / override / fraude)",
               1, 2, 8, height=60)
    r += 1
    _field_row(ws, r, "Vinculação a Achado (ID Achado se exceção gera deficiência)", 1, 2, 4)
    _field_row(ws, r, "Encaminhado para Tracker de Achados?", 5, 6, 8)
    _add_dv(ws, "F", r, r, SIM_NAO)

    return ws


# ════════════════════════════════════════════════════════════════════════════
# ABA 2 — Resultados por Item
# ════════════════════════════════════════════════════════════════════════════
def build_resultados(wb):
    ws = wb.create_sheet("Resultados por Item")
    ws.sheet_view.showGridLines = False

    N_COLS = 12

    ws.merge_cells("A1:L1")
    t = ws["A1"]
    t.value = "RESULTADOS POR ITEM TESTADO"
    t.fill  = _fill(VERDE_HEADER)
    t.font  = _font(bold=True, size=13)
    t.alignment = _center()
    t.border = _border_thin()
    ws.row_dimensions[1].height = 30

    _hint_merged(ws, 2,
        "Registre um item por linha. Preencha 'É Exceção?' antes de classificar a causa. "
        "Causa deve ser consistente com a definição de exceção pré-execução do Cabeçalho.",
        N_COLS)

    headers = [
        (1,  "Item #",                       8,  VERDE_HEADER),
        (2,  "Referência / ID do Item",      22,  VERDE_HEADER),
        (3,  "Data da Transação",            16,  VERDE_HEADER),
        (4,  "Valor (se aplicável)",         16,  VERDE_HEADER),
        (5,  "Atributo 1 Verificado",        32,  AZUL_SUB),
        (6,  "Atributo 2 Verificado",        32,  AZUL_SUB),
        (7,  "Atributo 3 Verificado",        32,  AZUL_SUB),
        (8,  "Evidência (referência/path)",  32,  AZUL_SUB),
        (9,  "Resultado do Item",            18,  VERDE_HEADER),
        (10, "É Exceção?",                  12,  "C00000"),
        (11, "Causa da Exceção",            32,  "C00000"),
        (12, "Comentário / Observação",     40,  VERDE_HEADER),
    ]

    for col, label, width, fhex in headers:
        cell = ws.cell(row=3, column=col, value=label)
        cell.fill = _fill(fhex)
        cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True)
        cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 36
    ws.freeze_panes = "A4"

    N = 100
    end = 4 + N - 1
    for row in range(4, end + 1):
        fill = _fill(CINZA_ALT) if row % 2 == 0 else _fill(BRANCO)
        for col in range(1, N_COLS + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = _border_thin()
            cell.alignment = _left() if col > 1 else _center()
            cell.font = Font(size=10, name="Arial", color="000000")
        ws.cell(row=row, column=1, value=row - 3)

    _add_dv(ws, "I", 4, end, RESULT_ITEM)
    _add_dv(ws, "J", 4, end, ["Sim", "Não"])
    _add_dv(ws, "K", 4, end, CAUSA_EXCEP)

    # linha de exemplo
    ex = [
        1, "TXN-2024-001", "01/03/2024", "R$ 45.200,00",
        "Aprovação nível 1 presente e dentro da alçada",
        "Aprovação nível 2 presente e dentro da alçada",
        "Aprovação executada antes do processamento",
        "ERP_ApprovalLog_Mar24.pdf — p. 3",
        "Aprovado", "Não", "", "",
    ]
    for col, val in enumerate(ex, 1):
        cell = ws.cell(row=4, column=col, value=val)
        cell.font = Font(size=10, name="Arial", color="595959", italic=True)

    return ws


# ════════════════════════════════════════════════════════════════════════════
# ABA 3 — _Ref_Amostragem
# ════════════════════════════════════════════════════════════════════════════
def build_ref_amostragem(wb):
    ws = wb.create_sheet("_Ref_Amostragem")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "REFERÊNCIA — TAMANHO DE AMOSTRA PARA TESTES DE CONTROLE (sample-size.md)"
    t.fill  = _fill(AZUL_HEADER)
    t.font  = _font(bold=True, size=11)
    t.alignment = _center()
    t.border = _border_thin()
    ws.row_dimensions[1].height = 28

    ws["A2"] = ("Fonte: _method-wiki/concepts/sample-size.md — referencial metodológico Cap. 8 + AICPA. "
                "Fórmula base: N = F / T   (F = fator de confiança; T = taxa tolerável como decimal)")
    ws["A2"].font = Font(size=9, name="Arial", color="595959", italic=True)
    ws.row_dimensions[2].height = 18

    r = 4

    # ── Tabela 1: Fatores de Confiança ────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    h = ws.cell(row=r, column=1, value="FATORES DE CONFIANÇA (F) — PARA CÁLCULO N = F / T")
    h.fill = _fill(AZUL_SUB); h.font = _font(bold=True); h.alignment = _center()
    h.border = _border_thin(); ws.row_dimensions[r].height = 22; r += 1

    fc_headers = ["Nível de Confiança", "Fator F", "Uso Típico"]
    fc_widths   = [26, 14, 50]
    for ci, (hdr, w) in enumerate(zip(fc_headers, fc_widths), 1):
        cell = ws.cell(row=r, column=ci, value=hdr)
        cell.fill = _fill(AZUL_SUB); cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True); cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[r].height = 28; r += 1

    fatores = [
        ("99%", "4,61", "Controles de alta criticidade / SOX material weakness"),
        ("95%", "3,00", "Padrão para assurance; controles significativos"),
        ("90%", "2,31", "Padrão interno; maioria dos controles operacionais"),
        ("87%", "2,00", "Simplificado para baixo risco (N arredondado = F/T)"),
        ("80%", "1,61", "Controles de baixa criticidade / procedimentos analíticos"),
        ("75%", "1,39", "Testes preliminares / planejamento"),
    ]
    for row_data in fatores:
        fill = _fill(CINZA_ALT) if r % 2 == 0 else _fill(BRANCO)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.fill = fill; cell.border = _border_thin()
            cell.alignment = _left(wrap=(ci == 3))
            cell.font = Font(size=10, name="Arial", color="000000",
                             bold=(ci in (1, 2)))
        ws.row_dimensions[r].height = 22; r += 1

    ws.merge_cells(f"A{r}:F{r}")
    ws.cell(row=r, column=1,
            value="Exemplo: 90% confiança, 5% tolerância → N = 2,31 / 0,05 = 47 itens").font = \
        Font(size=9, name="Arial", color="595959", italic=True)
    ws.row_dimensions[r].height = 18; r += 2

    # ── Tabela 2: Controles por Frequência ────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    h2 = ws.cell(row=r, column=1, value="AMOSTRAS POR FREQUÊNCIA DO CONTROLE")
    h2.fill = _fill(AZUL_SUB); h2.font = _font(bold=True); h2.alignment = _center()
    h2.border = _border_thin(); ws.row_dimensions[r].height = 22; r += 1

    freq_headers = ["Frequência", "Ocorrências/ano", "n (90% / 5%)",
                    "n (95% / 5%)", "Controle Automatizado", "Observação metodológica"]
    freq_widths  = [26, 18, 16, 16, 24, 52]
    for ci, (hdr, w) in enumerate(zip(freq_headers, freq_widths), 1):
        cell = ws.cell(row=r, column=ci, value=hdr)
        cell.fill = _fill(AZUL_SUB); cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True); cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[r].height = 36; r += 1

    # Valores canônicos extraídos de sample-size.md
    tabela_freq = [
        ("Diário (≥250 ocorrências)", "≥250",  "46",  "60",  "1–2",
         "Para automatizados: 1-2 instâncias se ITGCs efetivos. "
         "Falha em automatizado = achado imediato para toda a população."),
        ("Semanal (≈52 ocorrências)", "≈52",   "5–9", "9–15", "1–2",
         "Testar distribuição no período; evitar concentração em apenas um mês."),
        ("Quinzenal (≈24 ocorrências)", "≈24", "3–8", "5–10", "1–2",
         "Cobrir pelo menos dois trimestres distintos na amostra."),
        ("Mensal (12 ocorrências)",   "12",    "2–4", "4–6",  "1–2",
         "Para n=2: cobrir início e fim do período. Para n=4: distribuir pelos trimestres."),
        ("Trimestral (4 ocorrências)", "4",    "2",   "2–3",  "1",
         "Testar os 2 trimestres de maior risco; se baixo risco, 1 instância pode ser suficiente."),
        ("Semestral (2 ocorrências)", "2",     "2",   "2",    "1",
         "Testar ambas as ocorrências do período auditado."),
        ("Anual (1 ocorrência)",       "1",    "1",   "1",    "1",
         "Testar a única ocorrência; ampliar escopo se risco alto (ex.: revisão documental completa)."),
        ("Sob Demanda / Ad hoc",  "variável",  "46",  "60",   "1–2",
         "Tratar como diário se volume alto. Para volume baixo (<25), testar toda a população."),
    ]
    for row_data in tabela_freq:
        fill = _fill(CINZA_ALT) if r % 2 == 0 else _fill(BRANCO)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.fill = fill; cell.border = _border_thin()
            cell.alignment = _left(wrap=(ci == 6))
            cell.font = Font(size=10, name="Arial", color="000000",
                             bold=(ci == 1))
        ws.row_dimensions[r].height = 38; r += 1

    r += 1

    # ── Tabela 3: Dois Estágios ────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    h3 = ws.cell(row=r, column=1, value="TESTE SEQUENCIAL DOIS ESTÁGIOS (Two-Stage Sequential Sampling)")
    h3.fill = _fill(ROXO_HEADER); h3.font = _font(bold=True); h3.alignment = _center()
    h3.border = _border_thin(); ws.row_dimensions[r].height = 22; r += 1

    two_stage_headers = ["Estágio", "n Inicial", "Condição", "Decisão", "Observação"]
    two_stage_widths  = [16, 14, 35, 35, 42]
    for ci, (hdr, w) in enumerate(zip(two_stage_headers, two_stage_widths), 1):
        cell = ws.cell(row=r, column=ci, value=hdr)
        cell.fill = _fill(ROXO_HEADER); cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True); cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[r].height = 28; r += 1

    two_stage = [
        ("Estágio 1", "15–51*",
         "0 exceções",
         "Concluir: controle efetivo a 90% confiança",
         "n varia por tolerância: tol=10% → n=15; tol=5% → n=46; tol=7% → n=33"),
        ("Estágio 1", "15–51*",
         "1 exceção",
         "Ir para Estágio 2",
         "Não concluir ainda — evidência insuficiente para aceitar ou rejeitar"),
        ("Estágio 1", "15–51*",
         "2+ exceções",
         "Controle ineficaz — encerrar",
         "Documentar como deficiência; não expandir amostra"),
        ("Estágio 2", "+10 a +25",
         "0 exceções adicionais",
         "Concluir: controle efetivo (exceção do Estágio 1 = desvio isolado)",
         "Total n = Estágio1 + Estágio2; documentar análise qualitativa do desvio"),
        ("Estágio 2", "+10 a +25",
         "1+ exceção adicional",
         "Controle ineficaz — encerrar",
         "2+ exceções no total configuram padrão sistêmico"),
    ]
    for row_data in two_stage:
        fill = _fill(CINZA_ALT) if r % 2 == 0 else _fill(BRANCO)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.fill = fill; cell.border = _border_thin()
            cell.alignment = _left(wrap=True)
            cell.font = Font(size=10, name="Arial", color="000000",
                             bold=(ci == 1))
        ws.row_dimensions[r].height = 40; r += 1

    ws.merge_cells(f"A{r}:F{r}")
    ws.cell(row=r, column=1,
            value="* n do Estágio 1 = F / T  com F para 90%. Para tol.=10%: 2,31/0,10=23; "
                  "tol.=7%: 2,31/0,07=33; tol.=5%: 2,31/0,05=46. Ref: AICPA / referencial metodológico App. 8A.").font = \
        Font(size=9, name="Arial", color="595959", italic=True)
    ws.row_dimensions[r].height = 24; r += 2

    # ── Nota: Decisões pós-amostragem ─────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    h4 = ws.cell(row=r, column=1, value="DECISÕES PÓS-AMOSTRAGEM (sample-size.md §Post-sampling)")
    h4.fill = _fill(AZUL_SUB); h4.font = _font(bold=True); h4.alignment = _center()
    h4.border = _border_thin(); ws.row_dimensions[r].height = 22; r += 1

    notas = [
        ("0 exceções",
         "Suporta confiança planejada. Documentar como Efetivo."),
        ("1 exceção não planejada",
         "Opções: (a) dobrar N, (b) reduzir reliance, (c) remediar imediatamente + re-testar."),
        ("Múltiplas exceções",
         "Investigar causa sistêmica — assumir falha em toda a população até prova contrária."),
        ("Exceção pré-período",
         "Gaps de período requerem fechamento evidenciado até o fim do período auditado."),
        ("Controle automatizado com exceção",
         "Achado imediato para toda a população; verificar ITGCs e change management."),
    ]
    for situacao, acao in notas:
        fill = _fill(CINZA_ALT) if r % 2 == 0 else _fill(BRANCO)
        ws.cell(row=r, column=1, value=situacao).fill = fill
        ws.cell(row=r, column=1).font = Font(size=10, name="Arial", color="000000", bold=True)
        ws.cell(row=r, column=1).border = _border_thin()
        ws.cell(row=r, column=1).alignment = _left()
        ws.merge_cells(f"B{r}:F{r}")
        ws.cell(row=r, column=2, value=acao).fill = fill
        ws.cell(row=r, column=2).font = Font(size=10, name="Arial", color="000000")
        ws.cell(row=r, column=2).border = _border_thin()
        ws.cell(row=r, column=2).alignment = _left(wrap=True)
        ws.row_dimensions[r].height = 28; r += 1

    return ws


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════
def main():
    script_dir  = Path(__file__).parent
    repo_root   = script_dir.parent
    default_out = repo_root / "templates" / "planilhas" / "Papel_Teste_Controle_v3_template.xlsx"
    out_path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_out

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_cabecalho(wb)
    build_resultados(wb)
    build_ref_amostragem(wb)

    wb.save(out_path)
    print(f"Template gerado: {out_path}")


if __name__ == "__main__":
    main()
