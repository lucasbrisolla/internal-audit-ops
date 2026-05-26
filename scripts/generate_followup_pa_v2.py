"""
generate_followup_pa_v2.py
Follow-up de Plano de Ação — versão 3 (slim).

Removido v2→v3: Nível de Comunicação Requerido, Área do Responsável,
                Dias em Atraso, Verificado Por (auditor).
Resultado: 22→17 colunas na aba principal.

Fontes:
  states.yml                                                      → estados canônicos PA e achado
  _method-wiki/patterns/control-deficiency-severity.md           → severidade com rótulos completos,
                                                                    cascata de comunicação
  _method-wiki/checklists/audit-artifacts-definition-of-done.md → critério de conclusão verificável,
                                                                    tipo de evidência esperada

Abas:
  1. Follow-up de Plano de Ação  — lista mestre com severidade canônica e evidência tipificada
  2. Histórico de Verificação    — registro por ciclo
  3. Painel de Aging             — aging + status consolidado
  4. _Ref_SeveridadeEstados      — escala completa de severidade + estados canônicos + cascata

Uso:
    python3 scripts/generate_followup_pa_v2.py [saida.xlsx]
    Default: templates/planilhas/Followup_Plano_Acao_v3_template.xlsx
"""
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("openpyxl não instalado. Execute: pip install openpyxl")

# ── Paleta canônica ──────────────────────────────────────────────────────────
AZUL_HEADER  = "1F4E79"
AZUL_SUB     = "2F75B6"
VERDE_HEADER = "375623"
ROXO_HEADER  = "7030A0"
CINZA_TITULO = "D6DCE4"
BRANCO       = "FFFFFF"
CINZA_ALT    = "F2F2F2"
AMARELO_HINT = "FFF2CC"

# ── Helpers ──────────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=BRANCO, size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")

def _border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _add_dv(ws, col_letter, r_start, r_end, choices):
    formula = '"' + ",".join(choices) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.sqref = f"{col_letter}{r_start}:{col_letter}{r_end}"
    ws.add_data_validation(dv)

def _title_row(ws, text, last_col, fill_hex=AZUL_HEADER, size=13):
    lc = get_column_letter(last_col)
    ws.merge_cells(f"A1:{lc}1")
    t = ws["A1"]
    t.value = text
    t.fill  = _fill(fill_hex)
    t.font  = _font(bold=True, size=size)
    t.alignment = _center()
    t.border = _border_thin()
    ws.row_dimensions[1].height = 30

def _hint_row(ws, text, last_col, row=2):
    lc = get_column_letter(last_col)
    ws.merge_cells(f"A{row}:{lc}{row}")
    h = ws.cell(row=row, column=1, value=text)
    h.fill  = _fill(AMARELO_HINT)
    h.font  = Font(size=9, name="Arial", color="595959", italic=True)
    h.alignment = _left(wrap=True)
    h.border = _border_thin()
    ws.row_dimensions[row].height = 28
    for col in range(2, last_col + 1):
        ws.cell(row=row, column=col).border = _border_thin()

def _set_headers(ws, headers, row=3):
    for col, label, width, fill_hex in headers:
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = _fill(fill_hex)
        cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True)
        cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[row].height = 44

def _data_rows(ws, n_cols, start_row, n_rows):
    for row in range(start_row, start_row + n_rows):
        fill = _fill(CINZA_ALT) if row % 2 == 0 else _fill(BRANCO)
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = _border_thin()
            cell.alignment = _left()
            cell.font = Font(size=10, name="Arial", color="000000")

def _example_row(ws, values, row):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(size=10, name="Arial", color="595959", italic=True)


# ── Domínio canônico ─────────────────────────────────────────────────────────

# states.yml — plano de ação
ESTADOS_PA = ["aberto", "em_andamento", "implementado", "verificado", "vencido"]

# control-deficiency-severity.md — rótulos completos canônicos
SEVERIDADES = [
    "Crítico (Material Weakness)",
    "Alto (Significant Deficiency)",
    "Moderado (Deficiency)",
    "Baixo (Exception / Obs. de Melhoria)",
]

# Cores por status para highlight visual
STATUS_COLORS = {
    "aberto":       ("92D050", "000000"),
    "em_andamento": ("FFFF00", "000000"),
    "implementado": ("0070C0", BRANCO),
    "verificado":   ("375623", BRANCO),
    "vencido":      ("FF0000", BRANCO),
}

# Cores por severidade
SEV_COLORS = {
    "Crítico (Material Weakness)":          ("FF0000", BRANCO),
    "Alto (Significant Deficiency)":        ("FF9900", "000000"),
    "Moderado (Deficiency)":                ("FFFF00", "000000"),
    "Baixo (Exception / Obs. de Melhoria)": ("92D050", "000000"),
}

PROCESS_CODES = [
    "OTC","REV","P2P","RTR","H2R","FA","INV","TRE","TAX",
    "CAPEX","ITGC","UAM","JE","ELC","FRA","DRI","LCT","SCL","ESG",
]

RESULTADO_VER = [
    "Implementado — Efetivo",
    "Implementado — Parcial (risco residual remanescente)",
    "Não Implementado",
    "Implementação em Progresso",
    "Evidência Insuficiente — reverificar",
]

# Tipificação de evidência — DoD: tipo esperado definido antes da verificação
TIPO_EVIDENCIA = [
    "Configuração de sistema / parâmetro",
    "Policy / Procedimento atualizado",
    "Treinamento com lista de presença",
    "Relatório de controle operando",
    "Print screen / Log de sistema",
    "Aprovação formal / E-mail de aprovação",
    "Matriz de perfis / SoD atualizada",
    "Conciliação / Relatório de exceção",
    "Teste de redesenho pelo auditor",
    "Outro",
]

# Cascata de comunicação (severity.md)
NIVEL_COMUNICACAO = [
    "Responsável pelo Processo",
    "Nível acima do Responsável",
    "Diretoria / CFO",
    "Comitê de Auditoria / Board",
]

SIM_NAO = ["Sim", "Não", "N/A"]
N_ROWS = 120


# ════════════════════════════════════════════════════════════════════════════
# ABA 1 — Follow-up de Plano de Ação
# ════════════════════════════════════════════════════════════════════════════
def build_followup(wb):
    ws = wb.create_sheet("Follow-up de Plano de Ação")
    ws.sheet_view.showGridLines = False

    N_COLS = 17
    _title_row(ws, "FOLLOW-UP DE PLANO DE AÇÃO — AUDITORIA INTERNA  v3", N_COLS)
    _hint_row(ws,
        "Severidade com rótulo completo: ex. 'Alto (Significant Deficiency)'. "
        "Estado segue states.yml: aberto → em_andamento → implementado → verificado / vencido. "
        "Preencher 'Tipo de Evidência Esperada' antes de iniciar a verificação (DoD).",
        N_COLS)

    COR_ID    = AZUL_HEADER
    COR_ACAO  = AZUL_SUB
    COR_VERIF = VERDE_HEADER
    COR_SEV   = ROXO_HEADER

    headers = [
        # Identificação
        (1,  "ID Ação",                           10, COR_ID),
        (2,  "ID Achado Vinculado",               16, COR_ID),
        (3,  "Processo",                          16, COR_ID),
        (4,  "Título do Achado",                  40, COR_ID),
        # Severidade
        (5,  "Severidade do Achado",              30, COR_SEV),
        # Ação
        (6,  "Descrição da Ação",                 50, COR_ACAO),
        (7,  "Responsável (Owner)",               25, COR_ACAO),
        (8,  "Data-Alvo Original",                18, COR_ACAO),
        (9,  "Data-Alvo Revisada",                18, COR_ACAO),
        (10, "Critério de Conclusão Verificável", 44, COR_ACAO),
        (11, "Tipo de Evidência Esperada",        30, COR_ACAO),
        # Status
        (12, "Status Atual",                      20, COR_ID),
        (13, "Data da Última Atualização",        22, COR_ID),
        # Verificação
        (14, "Verificado pela Auditoria?",        22, COR_VERIF),
        (15, "Data da Verificação",               18, COR_VERIF),
        (16, "Resultado da Verificação",          35, COR_VERIF),
        (17, "Observações / Histórico",           40, COR_ID),
    ]
    _set_headers(ws, headers, row=3)
    ws.freeze_panes = "A4"

    end = 4 + N_ROWS - 1
    _data_rows(ws, N_COLS, 4, N_ROWS)

    _add_dv(ws, "C",  4, end, PROCESS_CODES)
    _add_dv(ws, "E",  4, end, SEVERIDADES)
    _add_dv(ws, "K",  4, end, TIPO_EVIDENCIA)
    _add_dv(ws, "L",  4, end, ESTADOS_PA)
    _add_dv(ws, "N",  4, end, SIM_NAO)
    _add_dv(ws, "P",  4, end, RESULTADO_VER)

    # linha de exemplo
    ex = [
        "PA-001", "ACH-001", "P2P",
        "Ausência de segregação de funções entre cadastro e pagamento",
        "Alto (Significant Deficiency)",
        "Redesenhar perfis SAP segregando as funções de cadastro, aprovação e pagamento "
        "conforme matriz SoD aprovada pelo CISO.",
        "Carlos Menezes",
        "30/06/2025", "",
        "Matriz de perfis SAP atualizada sem conflitos SoD + homologação formal pelo CISO.",
        "Matriz de perfis / SoD atualizada",
        "em_andamento", "15/04/2025",
        "Não", "", "",
        "Aguardando homologação do departamento de TI — prevista 20/05/2025.",
    ]
    for col, val in enumerate(ex, 1):
        cell = ws.cell(row=4, column=col, value=val)
        # highlight severidade e status
        if col == 5:
            fhex, fc = SEV_COLORS.get(val, (BRANCO, "000000"))
            cell.fill = _fill(fhex)
            cell.font = Font(size=10, name="Arial", bold=True, color=fc, italic=False)
        elif col == 12:
            fhex, fc = STATUS_COLORS.get(val, (BRANCO, "000000"))
            cell.fill = _fill(fhex)
            cell.font = Font(size=10, name="Arial", bold=True, color=fc, italic=False)
        else:
            cell.font = Font(size=10, name="Arial", color="595959", italic=True)

    return ws


# ════════════════════════════════════════════════════════════════════════════
# ABA 2 — Histórico de Verificação
# ════════════════════════════════════════════════════════════════════════════
def build_historico(wb):
    ws = wb.create_sheet("Histórico de Verificação")
    ws.sheet_view.showGridLines = False

    N_COLS = 10
    _title_row(ws, "HISTÓRICO DE VERIFICAÇÃO POR CICLO", N_COLS, fill_hex=VERDE_HEADER)
    _hint_row(ws,
        "Um registro por ID Ação por ciclo de verificação. "
        "Resultado deve referenciar a evidência analisada — não apenas declarar implementado.",
        N_COLS)

    headers = [
        (1,  "ID Ação",                     10, VERDE_HEADER),
        (2,  "ID Achado",                   14, VERDE_HEADER),
        (3,  "Ciclo de Acompanhamento",     22, VERDE_HEADER),
        (4,  "Data da Verificação",         18, VERDE_HEADER),
        (5,  "Auditor Verificador",         22, VERDE_HEADER),
        (6,  "Tipo de Evidência Analisada", 30, AZUL_SUB),
        (7,  "Evidência Analisada (ref.)",  40, AZUL_SUB),
        (8,  "Resultado",                   35, VERDE_HEADER),
        (9,  "Status Apurado",              20, VERDE_HEADER),
        (10, "Comentário / Próximo Passo",  45, VERDE_HEADER),
    ]
    _set_headers(ws, headers, row=3)
    ws.freeze_panes = "A4"

    end = 4 + N_ROWS - 1
    _data_rows(ws, N_COLS, 4, N_ROWS)

    _add_dv(ws, "F", 4, end, TIPO_EVIDENCIA)
    _add_dv(ws, "H", 4, end, RESULTADO_VER)
    _add_dv(ws, "I", 4, end, ESTADOS_PA)

    return ws


# ════════════════════════════════════════════════════════════════════════════
# ABA 3 — Painel de Aging
# ════════════════════════════════════════════════════════════════════════════
def build_painel_aging(wb):
    ws = wb.create_sheet("Painel de Aging")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "PAINEL DE AGING — PLANOS DE AÇÃO  v3"
    t.fill  = _fill(AZUL_HEADER); t.font = _font(bold=True, size=13)
    t.alignment = _center(); t.border = _border_thin()
    ws.row_dimensions[1].height = 30

    ws["A2"] = ("Referência para acompanhamento executivo. "
                "Atualizar a partir da aba 'Follow-up de Plano de Ação' periodicamente.")
    ws["A2"].font = Font(size=9, name="Arial", color="595959", italic=True)
    ws.row_dimensions[2].height = 18

    col_widths = [28, 14, 22, 22, 22, 22, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Tabela 1: Aging por faixa ─────────────────────────────────────────
    ws.merge_cells("A4:H4")
    hdr = ws.cell(row=4, column=1, value="RESUMO POR FAIXA DE ATRASO")
    hdr.fill = _fill(AZUL_SUB); hdr.font = _font(bold=True, size=11)
    hdr.alignment = _center(); hdr.border = _border_thin()

    aging_headers = [
        "Faixa de Atraso", "Qtd. Ações",
        "Crítico (MW)", "Alto (SD)", "Moderado (Def.)", "Baixo (Exc.)",
        "% do Total", "Escalonamento?",
    ]
    for ci, h in enumerate(aging_headers, 1):
        cell = ws.cell(row=5, column=ci, value=h)
        cell.fill = _fill(AZUL_HEADER); cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True); cell.border = _border_thin()
    ws.row_dimensions[5].height = 32

    faixas = [
        ("No prazo (0 dias)",      "0F9D58", "000000", "Não"),
        ("Atrasado 1–30 dias",     "FFFF00", "000000", "Crítico/Alto → Sim"),
        ("Atrasado 31–60 dias",    "FF9900", "000000", "Todos → Sim"),
        ("Atrasado 61–90 dias",    "FF6600", BRANCO,   "Todos → Sim + Comitê"),
        ("Atrasado >90 dias",      "FF0000", BRANCO,   "Todos → Board"),
        ("TOTAL",                  AZUL_HEADER, BRANCO, ""),
    ]
    for row_i, (label, fhex, fc, escal) in enumerate(faixas, 6):
        is_total = label == "TOTAL"
        for ci in range(1, 9):
            cell = ws.cell(row=row_i, column=ci)
            if ci == 1:
                cell.value = label
                cell.fill = _fill(fhex)
                cell.font = Font(bold=is_total, size=10, name="Arial", color=fc)
                cell.alignment = _left(wrap=False)
            elif ci == 8:
                cell.value = escal
                cell.fill = _fill(CINZA_ALT) if row_i % 2 == 0 else _fill(BRANCO)
                cell.font = Font(size=9, name="Arial", color="000000")
                cell.alignment = _left(wrap=True)
            else:
                cell.fill = _fill(CINZA_ALT) if row_i % 2 == 0 else _fill(BRANCO)
                cell.font = Font(bold=is_total, size=10, name="Arial", color="000000")
                cell.alignment = _center()
            cell.border = _border_thin()
        ws.row_dimensions[row_i].height = 22

    # ── Tabela 2: Status ──────────────────────────────────────────────────
    ws.merge_cells("A14:H14")
    hdr2 = ws.cell(row=14, column=1, value="RESUMO POR STATUS (states.yml)")
    hdr2.fill = _fill(VERDE_HEADER); hdr2.font = _font(bold=True, size=11)
    hdr2.alignment = _center(); hdr2.border = _border_thin()

    for ci in range(1, 9):
        cell = ws.cell(row=15, column=ci, value=aging_headers[ci-1])
        cell.fill = _fill(VERDE_HEADER); cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True); cell.border = _border_thin()
    ws.row_dimensions[15].height = 32

    for row_i, (status, (fhex, fc)) in enumerate(STATUS_COLORS.items(), 16):
        for ci in range(1, 9):
            cell = ws.cell(row=row_i, column=ci,
                           value=(status if ci == 1 else ""))
            if ci == 1:
                cell.fill = _fill(fhex)
                cell.font = Font(bold=True, size=10, name="Arial", color=fc)
                cell.alignment = _left(wrap=False)
            else:
                cell.fill = _fill(CINZA_ALT) if row_i % 2 == 0 else _fill(BRANCO)
                cell.font = Font(size=10, name="Arial", color="000000")
                cell.alignment = _center()
            cell.border = _border_thin()
        ws.row_dimensions[row_i].height = 22

    return ws


# ════════════════════════════════════════════════════════════════════════════
# ABA 4 — _Ref_SeveridadeEstados
# ════════════════════════════════════════════════════════════════════════════
def build_ref_sev_estados(wb):
    ws = wb.create_sheet("_Ref_SeveridadeEstados")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "REFERÊNCIA — SEVERIDADE CANÔNICA + ESTADOS PA + CASCATA DE COMUNICAÇÃO"
    t.fill  = _fill(AZUL_HEADER); t.font = _font(bold=True, size=11)
    t.alignment = _center(); t.border = _border_thin()
    ws.row_dimensions[1].height = 28

    ws["A2"] = ("Fonte: _method-wiki/patterns/control-deficiency-severity.md (referencial metodológico App.10A) "
                "+ IIA IPPF §7 + states.yml")
    ws["A2"].font = Font(size=9, name="Arial", color="595959", italic=True)

    col_widths = [32, 42, 22, 22, 28, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 4

    # ── Tabela 1: Escala de Severidade ────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    h = ws.cell(row=r, column=1, value="ESCALA DE SEVERIDADE (referencial metodológico App.10A + IIA IPPF §7)")
    h.fill = _fill(ROXO_HEADER); h.font = _font(bold=True); h.alignment = _center()
    h.border = _border_thin(); ws.row_dimensions[r].height = 22; r += 1

    sev_headers = ["Rótulo Canônico", "Definição", "Comunicação Mínima",
                   "Prazo Max. Remediação", "Documentação Obrigatória", "Exemplos"]
    for ci, hdr in enumerate(sev_headers, 1):
        cell = ws.cell(row=r, column=ci, value=hdr)
        cell.fill = _fill(ROXO_HEADER); cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True); cell.border = _border_thin()
    ws.row_dimensions[r].height = 28; r += 1

    severidade_table = [
        ("Crítico (Material Weakness)",
         "Razoavelmente possível que distorção material não seja prevenida/detectada.",
         "Board / Comitê de Auditoria",
         "<30 dias",
         "Carta de Deficiência + Plano de Ação imediato + acompanhamento mensal",
         "Restatement; fraude sênior; ausência total de controle chave"),
        ("Alto (Significant Deficiency)",
         "Menos severo que MW, mas importante para governança e gestão.",
         "Diretoria / CFO",
         "30–90 dias",
         "Relatório formal de achado + plano de ação + follow-up trimestral",
         "SoD ausente em processo material; ITGC crítico comprometido"),
        ("Moderado (Deficiency)",
         "Controle não opera como esperado, mas não atinge limiar de SD.",
         "Nível acima do Responsável",
         "90–180 dias",
         "Achado no relatório + plano de ação monitorado",
         "Exceção isolada recorrente; evidência insuficiente em controle relevante"),
        ("Baixo (Exception / Obs. de Melhoria)",
         "Desvio isolado sem evidência de padrão; baixo risco residual.",
         "Responsável pelo Processo",
         "180+ dias / próximo ciclo",
         "Observação de melhoria no relatório ou memo interno",
         "Erro pontual; procedimento não documentado mas controle opera"),
    ]
    sev_fill_hex = {
        "Crítico (Material Weakness)":          "FF0000",
        "Alto (Significant Deficiency)":        "FF9900",
        "Moderado (Deficiency)":                "FFFF00",
        "Baixo (Exception / Obs. de Melhoria)": "92D050",
    }
    for row_data in severidade_table:
        label = row_data[0]
        fhex = sev_fill_hex.get(label, BRANCO)
        fc = BRANCO if fhex == "FF0000" else "000000"
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            if ci == 1:
                cell.fill = _fill(fhex)
                cell.font = Font(bold=True, size=10, name="Arial", color=fc)
            else:
                cell.fill = _fill(CINZA_ALT) if r % 2 == 0 else _fill(BRANCO)
                cell.font = Font(size=10, name="Arial", color="000000")
            cell.alignment = _left(wrap=True); cell.border = _border_thin()
        ws.row_dimensions[r].height = 50; r += 1

    r += 1

    # ── Tabela 2: Estados do Plano de Ação ────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    h2 = ws.cell(row=r, column=1, value="ESTADOS DO PLANO DE AÇÃO (states.yml)")
    h2.fill = _fill(VERDE_HEADER); h2.font = _font(bold=True); h2.alignment = _center()
    h2.border = _border_thin(); ws.row_dimensions[r].height = 22; r += 1

    est_headers = ["Estado", "Significado", "Quem Atualiza",
                   "Evidência de Transição", "Próximo Estado", "Alerta"]
    for ci, hdr in enumerate(est_headers, 1):
        cell = ws.cell(row=r, column=ci, value=hdr)
        cell.fill = _fill(VERDE_HEADER); cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True); cell.border = _border_thin()
    ws.row_dimensions[r].height = 28; r += 1

    estados_table = [
        ("aberto",
         "Compromisso recebido; implementação não iniciada.",
         "Gestão", "Aceitação formal do plano de ação", "em_andamento",
         "Se data-alvo < 30 dias → monitorar semanalmente"),
        ("em_andamento",
         "Gestão relatou ações em curso.",
         "Gestão", "Evidência parcial ou update de progresso", "implementado",
         "Se sem update por >30 dias → solicitar posição"),
        ("implementado",
         "Gestão reportou conclusão; aguarda verificação da auditoria.",
         "Gestão → Auditoria", "Evidência de implementação entregue", "verificado",
         "Auditoria deve verificar em até 30 dias após declaração"),
        ("verificado",
         "Auditoria confirmou implementação efetiva.",
         "Auditoria", "WP de verificação documentado", "— (encerrado)",
         "Se implementado parcial → reabrir como novo achado ou manter em_andamento"),
        ("vencido",
         "Prazo expirou sem implementação; requer escalonamento.",
         "Sistema / Auditoria", "Data-alvo ultrapassada sem status=implementado",
         "escalonamento",
         "Crítico/Alto → escalona em <7 dias. Moderado/Baixo → em <30 dias"),
    ]
    for row_data in estados_table:
        estado = row_data[0]
        fhex, fc = STATUS_COLORS.get(estado, (BRANCO, "000000"))
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            if ci == 1:
                cell.fill = _fill(fhex)
                cell.font = Font(bold=True, size=10, name="Arial", color=fc)
            else:
                cell.fill = _fill(CINZA_ALT) if r % 2 == 0 else _fill(BRANCO)
                cell.font = Font(size=10, name="Arial", color="000000")
            cell.alignment = _left(wrap=True); cell.border = _border_thin()
        ws.row_dimensions[r].height = 44; r += 1

    r += 1

    # ── Tabela 3: Cascata de Comunicação ─────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    h3 = ws.cell(row=r, column=1,
                 value="CASCATA DE COMUNICAÇÃO DE DEFICIÊNCIAS (severity.md §Communication cascade)")
    h3.fill = _fill(ROXO_HEADER); h3.font = _font(bold=True); h3.alignment = _center()
    h3.border = _border_thin(); ws.row_dimensions[r].height = 22; r += 1

    cascata = [
        ("Baixo (Exception)",           "Responsável pelo Processo",
         "Memo ou observação no relatório"),
        ("Moderado (Deficiency)",        "Nível acima do Responsável",
         "Relatório de auditoria com achado formal"),
        ("Alto (Significant Deficiency)","Diretoria / CFO",
         "Relatório formal + reunião de alinhamento"),
        ("Crítico (Material Weakness)",  "Comitê de Auditoria / Board",
         "Carta de Deficiência Material + reunião de Board + timeline de remediação"),
    ]
    casc_headers = ["Severidade", "Destinatário Mínimo", "Veículo de Comunicação"]
    for ci, hdr in enumerate(casc_headers, 1):
        cell = ws.cell(row=r, column=ci, value=hdr)
        ws.merge_cells(f"C{r}:F{r}") if ci == 3 else None
        cell.fill = _fill(ROXO_HEADER); cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True); cell.border = _border_thin()
    ws.row_dimensions[r].height = 24; r += 1

    for sev, dest, veiculo in cascata:
        fhex = sev_fill_hex.get(sev.split(" (")[0] + " (" + sev.split("(")[1]
                                if "(" in sev else sev, BRANCO)
        # match parcial
        fhex = next((v for k, v in sev_fill_hex.items() if sev in k or k in sev), BRANCO)
        fc = BRANCO if fhex == "FF0000" else "000000"
        fill_data = _fill(CINZA_ALT) if r % 2 == 0 else _fill(BRANCO)

        ws.cell(row=r, column=1, value=sev).fill = _fill(fhex)
        ws.cell(row=r, column=1).font = Font(bold=True, size=10, name="Arial", color=fc)
        ws.cell(row=r, column=1).border = _border_thin()
        ws.cell(row=r, column=1).alignment = _left()

        ws.cell(row=r, column=2, value=dest).fill = fill_data
        ws.cell(row=r, column=2).font = Font(size=10, name="Arial", color="000000")
        ws.cell(row=r, column=2).border = _border_thin()
        ws.cell(row=r, column=2).alignment = _left()

        ws.merge_cells(f"C{r}:F{r}")
        ws.cell(row=r, column=3, value=veiculo).fill = fill_data
        ws.cell(row=r, column=3).font = Font(size=10, name="Arial", color="000000")
        ws.cell(row=r, column=3).border = _border_thin()
        ws.cell(row=r, column=3).alignment = _left(wrap=True)
        ws.row_dimensions[r].height = 30; r += 1

    return ws


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════
def main():
    script_dir  = Path(__file__).parent
    repo_root   = script_dir.parent
    default_out = repo_root / "templates" / "planilhas" / "Followup_Plano_Acao_v3_template.xlsx"
    out_path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_out

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_followup(wb)
    build_historico(wb)
    build_painel_aging(wb)
    build_ref_sev_estados(wb)

    wb.save(out_path)
    print(f"Template gerado: {out_path}")


if __name__ == "__main__":
    main()
