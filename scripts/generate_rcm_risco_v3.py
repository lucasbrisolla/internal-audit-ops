"""
generate_rcm_risco_v3.py
RCM + Matriz de Riscos — versão 3 slim (auditoria interna operacional).

Removido em relação à v2:
  RCM     — Asserção, Componente COSO, Dependência IPE/ITGC, Sistema/IPE, Comp. Testado?
  Risco   — Fator de Risco, Apetite, Velocidade, Persistência (4D ERM), Racional Residual,
             Ação Requerida/Prazo

Abas:
  1. RCM              — 14 colunas (risco → controle → resultado)
  2. Matriz de Riscos — 13 colunas (inerente 2D + controle + residual + status)
  3. _Ref_RiskScoring — heat map 4×4, escala, lógica residual
  4. _Ref_ControlTypes — tipos de controle canônicos

Uso:
    python3 scripts/generate_rcm_risco_v3.py [saida.xlsx]
    Default: templates/planilhas/RCM_Matriz_Riscos_v3_template.xlsx
"""
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("openpyxl não instalado. Execute: pip install openpyxl")

# ── Paleta canônica ──────────────────────────────────────────────────────────
AZUL_HEADER   = "1F4E79"
AZUL_SUB      = "2F75B6"
VERDE_HEADER  = "375623"
ROXO_HEADER   = "7030A0"
CINZA_TITULO  = "D6DCE4"
BRANCO        = "FFFFFF"
CINZA_ALT     = "F2F2F2"
AMARELO_HINT  = "FFF2CC"

# ── Helpers ──────────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=BRANCO, size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")

def _border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _title_row(ws, text, last_col, fill_hex=AZUL_HEADER, size=13):
    col_letter = get_column_letter(last_col)
    ws.merge_cells(f"A1:{col_letter}1")
    t = ws["A1"]
    t.value = text
    t.fill  = _fill(fill_hex)
    t.font  = _font(bold=True, size=size)
    t.alignment = _center()
    t.border = _border_thin()
    ws.row_dimensions[1].height = 30

def _hint_row(ws, text, last_col, row=2):
    col_letter = get_column_letter(last_col)
    ws.merge_cells(f"A{row}:{col_letter}{row}")
    h = ws.cell(row=row, column=1, value=text)
    h.fill  = _fill(AMARELO_HINT)
    h.font  = Font(size=9, name="Arial", color="595959", italic=True)
    h.alignment = _left(wrap=True)
    h.border = _border_thin()
    ws.row_dimensions[row].height = 24

def _add_dv(ws, col_letter, r_start, r_end, choices):
    formula = '"' + ",".join(choices) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.sqref = f"{col_letter}{r_start}:{col_letter}{r_end}"
    ws.add_data_validation(dv)

def _set_headers(ws, headers, row=3, default_fill=AZUL_SUB):
    for col, label, width, fill_hex in headers:
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = _fill(fill_hex)
        cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True)
        cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[row].height = 40

def _data_rows(ws, n_cols, start_row, n_rows):
    for row in range(start_row, start_row + n_rows):
        fill = _fill(CINZA_ALT) if row % 2 == 0 else _fill(BRANCO)
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = _border_thin()
            cell.alignment = _left()
            cell.font = Font(size=10, name="Arial", color="000000")

def _example_row(ws, values, row):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(size=10, name="Arial", color="595959", italic=True)


# ── Domínio canônico ─────────────────────────────────────────────────────────
# control-types-and-reliance.md §1
TIPOS_CONTROLE = [
    "Manual",
    "IT-dependent",
    "Application control (Automatizado)",
    "Híbrido",
]
PREV_DET = ["Preventivo", "Detectivo", "Preventivo/Detectivo"]

FREQUENCIAS = [
    "Diário", "Semanal", "Quinzenal", "Mensal",
    "Trimestral", "Semestral", "Anual", "Sob Demanda",
]

# risk-scoring-foundations.md — escala 4×4
IMPACTO_4 = [
    "Insignificante (1) — <0,1% receita",
    "Baixo (2) — 0,1–0,5% receita",
    "Moderado (3) — 0,5–2% receita",
    "Significativo (4) — >2% receita",
]
PROB_4 = [
    "Improvável (1) — <1x em 3 anos",
    "Possível (2) — 1–2x/ano",
    "Provável (3) — mensal/trimestral",
    "Quase Certo (4) — semanal/contínuo",
]
CATEG_RISCO = [
    "Operacional", "Financeiro", "Conformidade/Regulatório",
    "Estratégico", "Tecnológico", "Reputacional", "Fraude",
]

# Bandas de score (heat map 4×4 assimétrico — risk-scoring-foundations.md)
# scores possíveis: 1-16; bandas: 1-3=Baixo, 4-9=Moderado, 10-12=Alto, 13-16=Crítico
SCORE_BANDS = {
    (1,1):("Baixo",1,"92D050"),    (1,2):("Baixo",2,"92D050"),
    (1,3):("Moderado",3,"FFFF00"), (1,4):("Moderado",4,"FFFF00"),
    (2,1):("Baixo",2,"92D050"),    (2,2):("Moderado",4,"FFFF00"),
    (2,3):("Moderado",6,"FFFF00"), (2,4):("Alto",8,"FF9900"),
    (3,1):("Moderado",3,"FFFF00"), (3,2):("Moderado",6,"FFFF00"),
    (3,3):("Alto",9,"FF9900"),     (3,4):("Alto",12,"FF9900"),
    (4,1):("Moderado",4,"FFFF00"), (4,2):("Alto",8,"FF9900"),
    (4,3):("Alto",12,"FF9900"),    (4,4):("Crítico",16,"FF0000"),
}

RESULT_CTRL  = [
    "Satisfatório (≥0,90)",
    "Satisfatório com ressalvas (0,75–0,89)",
    "Requer melhoria significativa (0,60–0,74)",
    "Insatisfatório (<0,60)",
    "Não avaliado",
]
RESULT_EFETIV = ["Efetivo", "Inefetivo", "Em andamento", "Não testado"]
PROCESS_CODES = [
    "FEC","FAT","REC","CAD-CLI","ENT","DEV",
    "CAD-AT","AQU","PAT","ALI","INV","TKL",
]

N_DATA = 150


# ════════════════════════════════════════════════════════════════════════════
# ABA 1 — RCM
# ════════════════════════════════════════════════════════════════════════════
def build_rcm(wb):
    ws = wb.create_sheet("RCM - Matriz de Controle")
    ws.sheet_view.showGridLines = False

    # 14 colunas: 4 risco + 3 identificação controle + 5 atributos + 2 resultado
    N_COLS = 14
    COR_RISCO   = AZUL_HEADER
    COR_CTRL_ID = AZUL_SUB
    COR_ATR     = "2E75B6"
    COR_RESULT  = VERDE_HEADER

    _title_row(ws, "RISK AND CONTROL MATRIX (RCM) — v3", N_COLS)
    _hint_row(ws,
        "Um controle por linha. Tipo: Manual | IT-dependent | Application control | Híbrido. "
        "Resultado Efetividade: Efetivo | Inefetivo | Em andamento | Não testado.",
        N_COLS)

    headers = [
        # Risco
        (1,  "Cód. Risco",           12, COR_RISCO),
        (2,  "Processo (code)",       14, COR_RISCO),
        (3,  "Subprocesso",          22, COR_RISCO),
        (4,  "WCGW",                 42, COR_RISCO),
        # Identificação do controle
        (5,  "Cód. Controle",        14, COR_CTRL_ID),
        (6,  "Título do Controle",   30, COR_CTRL_ID),
        (7,  "Descrição Detalhada",  52, COR_CTRL_ID),
        # Atributos
        (8,  "Frequência",           16, COR_ATR),
        (9,  "Tipo de Controle",     26, COR_ATR),
        (10, "Natureza",             18, COR_ATR),
        (11, "Owner (Executor)",     22, COR_ATR),
        (12, "Evidência Requerida",  38, COR_ATR),
        (13, "Critério de Falha",    38, COR_ATR),
        # Resultado
        (14, "Resultado Efetividade", 22, COR_RESULT),
    ]
    _set_headers(ws, headers, row=3)
    ws.freeze_panes = "A4"

    _data_rows(ws, N_COLS, 4, N_DATA)
    end = 4 + N_DATA - 1

    _add_dv(ws, "B", 4, end, PROCESS_CODES)
    _add_dv(ws, "H", 4, end, FREQUENCIAS)
    _add_dv(ws, "I", 4, end, TIPOS_CONTROLE)
    _add_dv(ws, "J", 4, end, PREV_DET)
    _add_dv(ws, "N", 4, end, RESULT_EFETIV)

    _example_row(ws, [
        "R-FEC-01", "FEC", "Fechamento de Contrato",
        "Contrato encerrado sem checklist completo — itens pendentes não identificados",
        "C-FEC-01", "Checklist de fechamento de contrato",
        "Responsável de contratos verifica, antes do fechamento no sistema operacional, "
        "todos os itens do checklist padrão (devolução de veículo, quitação de faturas, "
        "retirada de rastreador). Evidência: checklist assinado com data.",
        "Sob Demanda", "Manual", "Preventivo",
        "Coordenador de Contratos",
        "Checklist de fechamento assinado e datado, arquivado no SharePoint",
        "Fechamento registrado no sistema sem checklist aprovado ou com item em aberto",
        "Não testado",
    ], row=4)

    return ws


# ════════════════════════════════════════════════════════════════════════════
# ABA 2 — Matriz de Riscos
# ════════════════════════════════════════════════════════════════════════════
def build_risco(wb):
    ws = wb.create_sheet("Matriz de Riscos")
    ws.sheet_view.showGridLines = False

    # 13 colunas: 6 identificação + 3 inerente + 3 controle + 1 residual status
    N_COLS = 13
    _title_row(ws, "MATRIZ DE AVALIAÇÃO DE RISCOS — v3 (Heat Map 4×4)", N_COLS)
    _hint_row(ws,
        "Score Inerente = heat map 4×4 assimétrico (Impacto × Probabilidade). "
        "Residual: Satisfatório ≥0,90 → -2 níveis; 0,75–0,89 → -1 nível; <0,75 → sem redução. "
        "Ver _Ref_RiskScoring.",
        N_COLS)

    COR_BASE  = AZUL_HEADER
    COR_INER  = AZUL_SUB
    COR_CTRL  = "2E75B6"
    COR_RESID = "BF8F00"

    grupos = [
        (1,  6,  "IDENTIFICAÇÃO",       COR_BASE),
        (7,  9,  "RISCO INERENTE",      COR_INER),
        (10, 12, "CONTROLE MITIGADOR",  COR_CTRL),
        (13, 13, "RESIDUAL",            COR_RESID),
    ]
    for c_start, c_end, label, fhex in grupos:
        ws.merge_cells(f"{get_column_letter(c_start)}3:{get_column_letter(c_end)}3")
        c = ws.cell(row=3, column=c_start, value=label)
        c.fill = _fill(fhex)
        c.font = _font(bold=True, size=10)
        c.alignment = _center()
        c.border = _border_thin()
        for col in range(c_start + 1, c_end + 1):
            ws.cell(row=3, column=col).border = _border_thin()
    ws.row_dimensions[3].height = 22

    headers = [
        # Identificação
        (1,  "Cód. Risco",            12, COR_BASE),
        (2,  "Processo (code)",        14, COR_BASE),
        (3,  "Subprocesso",           22, COR_BASE),
        (4,  "Risco",                 42, COR_BASE),
        (5,  "Categoria",             20, COR_BASE),
        (6,  "Owner do Risco",        22, COR_BASE),
        # Inerente — 2 dimensões + score
        (7,  "Impacto",               26, COR_INER),
        (8,  "Probabilidade",         26, COR_INER),
        (9,  "Score Inerente",        16, COR_INER),
        # Controle
        (10, "Controles Mitigadores", 40, COR_CTRL),
        (11, "Resultado do Controle", 26, COR_CTRL),
        (12, "Observação / Lacuna",   36, COR_CTRL),
        # Residual
        (13, "Nível Residual / Status", 22, COR_RESID),
    ]
    _set_headers(ws, headers, row=4)
    ws.freeze_panes = "A5"

    end = 5 + N_DATA - 1
    _data_rows(ws, N_COLS, 5, N_DATA)

    _add_dv(ws, "B", 5, end, PROCESS_CODES)
    _add_dv(ws, "E", 5, end, CATEG_RISCO)
    _add_dv(ws, "G", 5, end, IMPACTO_4)
    _add_dv(ws, "H", 5, end, PROB_4)
    _add_dv(ws, "K", 5, end, RESULT_CTRL)
    _add_dv(ws, "M", 5, end, ["Baixo","Moderado","Alto","Crítico"])

    _example_row(ws, [
        "R-FEC-01", "FEC", "Fechamento de Contrato",
        "Contrato encerrado sem checklist completo — itens pendentes não detectados",
        "Operacional", "Coordenador de Contratos",
        "Significativo (4) — >2% receita", "Provável (3) — mensal/trimestral",
        "Alto (12)",
        "Checklist de fechamento manual (C-FEC-01); revisão 4-olhos semanal (C-FEC-02)",
        "Satisfatório com ressalvas (0,75–0,89)",
        "Checklist não cobre devolução de itens opcionais — lacuna de desenho",
        "Moderado",
    ], row=5)

    return ws


# ════════════════════════════════════════════════════════════════════════════
# ABA 3 — _Ref_RiskScoring
# ════════════════════════════════════════════════════════════════════════════
def build_ref_risk_scoring(wb):
    ws = wb.create_sheet("_Ref_RiskScoring")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "REFERÊNCIA — SCORING DE RISCO (risk-scoring-foundations.md)"
    t.fill  = _fill(AZUL_HEADER)
    t.font  = _font(bold=True, size=11)
    t.alignment = _center()
    t.border = _border_thin()
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Fonte: _method-wiki/concepts/risk-scoring-foundations.md — COSO ERM 2017"
    ws["A2"].font = Font(size=9, name="Arial", color="595959", italic=True)

    r = 4

    # ── Tabela 1: Heat Map 4×4 ───────────────────────────────────────────
    ws.merge_cells(f"A{r}:H{r}")
    h = ws.cell(row=r, column=1, value="HEAT MAP 4×4 ASSIMÉTRICO — SCORE DE RISCO INERENTE")
    h.fill = _fill(AZUL_SUB); h.font = _font(bold=True); h.alignment = _center()
    h.border = _border_thin(); ws.row_dimensions[r].height = 22; r += 1

    # cabeçalho prob
    ws.cell(row=r, column=1, value="Impacto \\ Probabilidade").fill = _fill(CINZA_TITULO)
    ws.cell(row=r, column=1).font = Font(bold=True, size=10, name="Arial", color="000000")
    ws.cell(row=r, column=1).alignment = _center(wrap=True)
    ws.cell(row=r, column=1).border = _border_thin()
    ws.column_dimensions["A"].width = 28

    prob_labels = [
        "Improvável (1)\n<1x em 3 anos",
        "Possível (2)\n1–2x/ano",
        "Provável (3)\nmensal/trim.",
        "Quase Certo (4)\nsemanal/cont.",
    ]
    for j, lp in enumerate(prob_labels, 2):
        c = ws.cell(row=r, column=j, value=lp)
        c.fill = _fill(AZUL_SUB); c.font = _font(bold=True, size=9)
        c.alignment = _center(wrap=True); c.border = _border_thin()
        ws.column_dimensions[get_column_letter(j)].width = 20
    ws.row_dimensions[r].height = 36; r += 1

    imp_labels = [
        "Insignificante (1)\n<0,1% receita",
        "Baixo (2)\n0,1–0,5%",
        "Moderado (3)\n0,5–2%",
        "Significativo (4)\n>2% receita",
    ]
    for i, li in enumerate(imp_labels, 1):
        c0 = ws.cell(row=r, column=1, value=li)
        c0.fill = _fill(CINZA_TITULO)
        c0.font = Font(bold=True, size=9, name="Arial", color="000000")
        c0.alignment = _left(wrap=True); c0.border = _border_thin()
        ws.row_dimensions[r].height = 36
        for j in range(1, 5):
            label, score, fhex = SCORE_BANDS[(i, j)]
            fc = BRANCO if fhex == "FF0000" else "000000"
            cell = ws.cell(row=r, column=j + 1, value=f"{label}\n({score})")
            cell.fill = _fill(fhex)
            cell.font = Font(bold=True, size=10, name="Arial", color=fc)
            cell.alignment = _center(wrap=True); cell.border = _border_thin()
        r += 1

    r += 1
    # Bandas
    ws.merge_cells(f"A{r}:H{r}")
    ws.cell(row=r, column=1,
            value="Bandas: 1–3 = Baixo  |  4–9 = Moderado  |  10–12 = Alto  |  13–16 = Crítico").font = \
        Font(size=9, name="Arial", color="595959", italic=True, bold=True)
    ws.row_dimensions[r].height = 18; r += 2

    # ── Tabela 2: Escala de Impacto ───────────────────────────────────────
    ws.merge_cells(f"A{r}:H{r}")
    h2 = ws.cell(row=r, column=1, value="ESCALA DE IMPACTO — DIMENSÃO FINANCEIRA (primária)")
    h2.fill = _fill(AZUL_SUB); h2.font = _font(bold=True); h2.alignment = _center()
    h2.border = _border_thin(); ws.row_dimensions[r].height = 22; r += 1

    imp_table = [
        ("1 — Insignificante", "<0,1% da receita líquida", "Sem impacto perceptível"),
        ("2 — Baixo",          "0,1–0,5% da receita",      "Resolúvel internamente"),
        ("3 — Moderado",       "0,5–2,0% da receita",      "Atenção da gestão"),
        ("4 — Significativo",  ">2,0% da receita",         "Escalonamento obrigatório"),
    ]
    for row_data in imp_table:
        fill = _fill(CINZA_ALT) if r % 2 == 0 else _fill(BRANCO)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.fill = fill; cell.border = _border_thin()
            cell.alignment = _left(); cell.font = Font(size=10, name="Arial", color="000000")
        r += 1

    ws.merge_cells(f"A{r}:H{r}")
    ws.cell(row=r, column=1,
            value="Avaliar também: impacto operacional, regulatório/reputacional. Usar a dimensão mais alta.").font = \
        Font(size=9, name="Arial", color="595959", italic=True)
    ws.row_dimensions[r].height = 18; r += 2

    # ── Tabela 3: Lógica de Risco Residual ────────────────────────────────
    ws.merge_cells(f"A{r}:H{r}")
    h3 = ws.cell(row=r, column=1, value="LÓGICA DE RISCO RESIDUAL — REDUÇÃO POR EFETIVIDADE DO CONTROLE")
    h3.fill = _fill(ROXO_HEADER); h3.font = _font(bold=True); h3.alignment = _center()
    h3.border = _border_thin(); ws.row_dimensions[r].height = 22; r += 1

    residual_headers = ["Efetividade do Controle", "Score", "Redução de Nível",
                        "Regras Especiais"]
    for ci, hdr in enumerate(residual_headers, 1):
        cell = ws.cell(row=r, column=ci, value=hdr)
        cell.fill = _fill(ROXO_HEADER); cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True); cell.border = _border_thin()
    ws.row_dimensions[r].height = 28; r += 1

    residual_table = [
        ("Satisfatório",              "≥0,90", "-2 níveis",
         "RI Crítico + Satisfatório = máximo Alto (não Baixo)"),
        ("Satisfatório com ressalvas", "0,75–0,89", "-1 nível",
         "Para fraude: máximo -1 nível independente do score"),
        ("Requer melhoria significativa", "0,60–0,74", "Sem redução",
         "Manter RI como base para RR"),
        ("Insatisfatório",            "<0,60",  "Sem redução",
         "RR = RI; gera achado obrigatório se RR Alto ou Crítico"),
        ("Não avaliado",              "—",      "Sem redução",
         "Tratar RR = RI até avaliação concluída"),
    ]
    for row_data in residual_table:
        fill = _fill(CINZA_ALT) if r % 2 == 0 else _fill(BRANCO)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.fill = fill; cell.border = _border_thin()
            cell.alignment = _left(wrap=ci == 4)
            cell.font = Font(size=10, name="Arial", color="000000")
        ws.row_dimensions[r].height = 30; r += 1

    ws.merge_cells(f"A{r}:H{r}")
    ws.cell(row=r, column=1,
            value="Regra geral: usar apenas o melhor controle (não somar). Mínimo residual = Baixo, exceto RI Crítico.").font = \
        Font(size=9, name="Arial", color="595959", italic=True)
    ws.row_dimensions[r].height = 18

    ws.column_dimensions["D"].width = 55
    return ws


# ════════════════════════════════════════════════════════════════════════════
# ABA 4 — _Ref_ControlTypes
# ════════════════════════════════════════════════════════════════════════════
def build_ref_control_types(wb):
    ws = wb.create_sheet("_Ref_ControlTypes")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "REFERÊNCIA — TIPOS DE CONTROLE E IMPLICAÇÕES (control-types-and-reliance.md)"
    t.fill  = _fill(AZUL_HEADER)
    t.font  = _font(bold=True, size=11)
    t.alignment = _center()
    t.border = _border_thin()
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Fonte: _method-wiki/concepts/control-types-and-reliance.md §1–§4"
    ws["A2"].font = Font(size=9, name="Arial", color="595959", italic=True)

    headers = ["Tipo", "Definição", "Dependência IPE/ITGC?",
               "Tamanho de Amostra", "Risco Principal", "Obs. de Documentação"]
    widths  = [30, 50, 22, 28, 35, 42]
    r = 3
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=r, column=ci, value=h)
        cell.fill = _fill(AZUL_SUB); cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True); cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[r].height = 28; r += 1

    tipos_table = [
        ("Manual",
         "Executado por pessoa sem dependência de sistema para operar.",
         "Não — mas pode usar outputs de sistema como input",
         "Amostra representativa (ver _Ref_Amostragem)",
         "Variabilidade humana; substituição de executor sem comunicação",
         "Documentar: quem executa, quando, como inicia, critério de revisão"),
        ("IT-dependent",
         "Executado por pessoa, mas depende de relatório/extração gerada pelo sistema.",
         "Sim — testar IPE + confiabilidade da extração",
         "Mesma lógica do Manual + teste de IPE",
         "Relatório alterado ou com parâmetros incorretos invalida o controle",
         "Identificar sistema, nome do relatório, parâmetros, responsável pela extração"),
        ("Application control (Automatizado)",
         "Lógica parametrizada no sistema — executa sem intervenção humana.",
         "Sim — depende de ITGCs efetivos (acesso, change mgmt, operações)",
         "1–2 instâncias se ITGCs efetivos; reverter para transacional se ITGCs fracos",
         "Erro de programação ou mudança não autorizada compromete toda a população",
         "Documentar: sistema, parâmetro de configuração, versão, data de last change"),
        ("Híbrido",
         "Combinação de lógica automatizada + revisão humana (ex.: aprovação em sistema + "
         "julgamento manual do revisor).",
         "Sim — testar a parte automatizada como Application control",
         "Testar cada componente separadamente; amostra para a parte manual",
         "Falha no componente manual invalida a efetividade mesmo com automação sadia",
         "Distinguir claramente o que é automatizado e o que depende de julgamento humano"),
    ]

    for row_data in tipos_table:
        fill = _fill(CINZA_ALT) if r % 2 == 0 else _fill(BRANCO)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.fill = fill; cell.border = _border_thin()
            cell.alignment = _left(wrap=True)
            cell.font = Font(size=10, name="Arial", color="000000",
                             bold=(ci == 1))
        ws.row_dimensions[r].height = 55; r += 1

    r += 1
    # Compensatório
    ws.merge_cells(f"A{r}:F{r}")
    hc = ws.cell(row=r, column=1, value="CONTROLES COMPENSATÓRIOS — Requisitos para Reliance")
    hc.fill = _fill(ROXO_HEADER); hc.font = _font(bold=True)
    hc.alignment = _left(); hc.border = _border_thin()
    ws.row_dimensions[r].height = 22; r += 1

    comp_items = [
        "1. Opera com precisão suficiente para cobrir o mesmo risco do controle primário falhado.",
        "2. Testado para efetividade operacional (não presumido efetivo).",
        "3. Sem histórico de falhas no período auditado.",
        "4. Apenas reduz (não elimina) o risco residual; nunca cancela uma deficiência de desenho.",
    ]
    for item in comp_items:
        cell = ws.cell(row=r, column=1, value=item)
        ws.merge_cells(f"A{r}:F{r}")
        cell.fill = _fill(CINZA_ALT if r % 2 == 0 else BRANCO)
        cell.font = Font(size=10, name="Arial", color="000000")
        cell.alignment = _left(); cell.border = _border_thin()
        ws.row_dimensions[r].height = 20; r += 1

    return ws


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════
def main():
    script_dir  = Path(__file__).parent
    repo_root   = script_dir.parent
    default_out = repo_root / "templates" / "planilhas" / "RCM_Matriz_Riscos_v3_template.xlsx"
    out_path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_out

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_rcm(wb)
    build_risco(wb)
    build_ref_risk_scoring(wb)
    build_ref_control_types(wb)

    wb.save(out_path)
    print(f"Template gerado: {out_path}")


if __name__ == "__main__":
    main()
