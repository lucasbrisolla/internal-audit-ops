"""
generate_plano_anual_v3.py
Plano Anual de Auditoria Interna — versão 3 slim (auditoria interna operacional).

Removido em relação à v2:
  Universo  — Velocidade, Persistência (4D ERM), Complexidade (subjetivo), Esforço (horas)
  Plano     — Nível de Risco (duplica Universo)
  Ref       — critério "Velocidade de Impacto" (15%); pesos redistribuídos

Abas:
  1. Universo Auditável   — 16 colunas (scoring 2D + planejamento)
  2. Plano Anual          — 21 colunas (calendário trimestral + status)
  3. Resumo Executivo     — painel sintético para comitê
  4. _Ref_Critérios       — critérios de priorização + heat map + estados

Uso:
    python3 scripts/generate_plano_anual_v3.py [saida.xlsx]
    Default: templates/planilhas/Plano_Anual_Auditoria_v3_template.xlsx
"""
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("openpyxl não instalado. Execute: pip install openpyxl")

# ── Paleta canônica ──────────────────────────────────────────────────────────
AZUL_HEADER  = "1F4E79"
AZUL_SUB     = "2F75B6"
VERDE_HEADER = "375623"
ROXO_HEADER  = "7030A0"
CINZA_TITULO = "D6DCE4"
BRANCO       = "FFFFFF"
CINZA_ALT    = "F2F2F2"
AMARELO_HINT = "FFF2CC"

# ── Helpers ──────────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=BRANCO, size=10):
    return Font(bold=bold, color=color, size=size, name="Arial")

def _border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def _left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

def _add_dv(ws, col_letter, r_start, r_end, choices):
    formula = '"' + ",".join(choices) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.sqref = f"{col_letter}{r_start}:{col_letter}{r_end}"
    ws.add_data_validation(dv)

def _title_row(ws, text, last_col, fill_hex=AZUL_HEADER, size=13):
    lc = get_column_letter(last_col)
    ws.merge_cells(f"A1:{lc}1")
    t = ws["A1"]
    t.value = text
    t.fill  = _fill(fill_hex)
    t.font  = _font(bold=True, size=size)
    t.alignment = _center()
    t.border = _border_thin()
    ws.row_dimensions[1].height = 30

def _hint_row(ws, text, last_col, row=2):
    lc = get_column_letter(last_col)
    ws.merge_cells(f"A{row}:{lc}{row}")
    h = ws.cell(row=row, column=1, value=text)
    h.fill  = _fill(AMARELO_HINT)
    h.font  = Font(size=9, name="Arial", color="595959", italic=True)
    h.alignment = _left(wrap=True)
    h.border = _border_thin()
    ws.row_dimensions[row].height = 28
    for col in range(2, last_col + 1):
        ws.cell(row=row, column=col).border = _border_thin()

def _set_headers(ws, headers, row=3):
    for col, label, width, fill_hex in headers:
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = _fill(fill_hex)
        cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True)
        cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[row].height = 44

def _data_rows(ws, n_cols, start_row, n_rows):
    for row in range(start_row, start_row + n_rows):
        fill = _fill(CINZA_ALT) if row % 2 == 0 else _fill(BRANCO)
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = _border_thin()
            cell.alignment = _left()
            cell.font = Font(size=10, name="Arial", color="000000")

def _example_row(ws, values, row):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(size=10, name="Arial", color="595959", italic=True)


# ── Domínio canônico ─────────────────────────────────────────────────────────
PROCESS_CODES = [
    "FEC","FAT","REC","CAD-CLI","ENT","DEV",
    "CAD-AT","AQU","PAT","ALI","INV","TKL",
]

# risk-scoring-foundations.md — escala 4×4
IMPACTO_4 = [
    "Insignificante (1) — <0,1% receita",
    "Baixo (2) — 0,1–0,5% receita",
    "Moderado (3) — 0,5–2% receita",
    "Significativo (4) — >2% receita",
]
PROB_4 = [
    "Improvável (1) — <1x em 3 anos",
    "Possível (2) — 1–2x/ano",
    "Provável (3) — mensal/trimestral",
    "Quase Certo (4) — semanal/contínuo",
]

# Score bands para lookup manual
SCORE_BANDS = {
    (1,1):("Baixo",1),    (1,2):("Baixo",2),
    (1,3):("Moderado",3), (1,4):("Moderado",4),
    (2,1):("Baixo",2),    (2,2):("Moderado",4),
    (2,3):("Moderado",6), (2,4):("Alto",8),
    (3,1):("Moderado",3), (3,2):("Moderado",6),
    (3,3):("Alto",9),     (3,4):("Alto",12),
    (4,1):("Moderado",4), (4,2):("Alto",8),
    (4,3):("Alto",12),    (4,4):("Crítico",16),
}

NIVEL_RISCO  = ["Baixo", "Moderado", "Alto", "Crítico"]

# states.yml — engagement
STATUS_ENGAGEMENT = [
    "planejamento",
    "execução",
    "revisão_técnica",
    "comunicação",
    "monitoramento",
    "arquivado",
    "adiado",
    "cancelado",
    "não_iniciado",
]

PRIORIDADE = ["1 - Crítico", "2 - Alto", "3 - Moderado", "4 - Baixo"]

TIPO_ENGAGEMENT = [
    "Assurance — Controles Internos",
    "Assurance — Conformidade/Regulatório",
    "Assurance — Operacional",
    "Assurance — TI / ITGC",
    "Assurance — Financeiro",
    "Consulting / Advisory",
    "Investigação",
    "Follow-up",
]

MESES = ["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"]
N_ROWS = 60


# ════════════════════════════════════════════════════════════════════════════
# ABA 1 — Universo Auditável
# ════════════════════════════════════════════════════════════════════════════
def build_universo(wb):
    ws = wb.create_sheet("Universo Auditável")
    ws.sheet_view.showGridLines = False

    # 16 colunas: 7 identificação + 4 scoring + 5 planejamento
    N_COLS = 16
    _title_row(ws, "UNIVERSO AUDITÁVEL — AVALIAÇÃO DE RISCO E PRIORIZAÇÃO  v3", N_COLS)
    _hint_row(ws,
        "Score Inerente = heat map 4×4 (Impacto × Probabilidade). "
        "Bandas: 1–3=Baixo | 4–9=Moderado | 10–12=Alto | 13–16=Crítico. "
        "Ver _Ref_Critérios.",
        N_COLS)

    COR_ID    = AZUL_HEADER
    COR_RISCO = AZUL_SUB
    COR_PLAN  = VERDE_HEADER

    headers = [
        # Identificação
        (1,  "ID",                       8,  COR_ID),
        (2,  "Área / Processo",         30,  COR_ID),
        (3,  "Código do Processo",      16,  COR_ID),
        (4,  "Subprocesso / Tema",      28,  COR_ID),
        (5,  "Tipo de Engagement",      30,  COR_ID),
        (6,  "Última Auditoria (ano)",  18,  COR_ID),
        (7,  "Owner do Processo",       22,  COR_ID),
        # Scoring de risco — 2 dimensões + score + nível
        (8,  "Impacto Inerente",        26,  COR_RISCO),
        (9,  "Probabilidade",           26,  COR_RISCO),
        (10, "Score Inerente (1–16)",   18,  COR_RISCO),
        (11, "Nível de Risco",          16,  COR_RISCO),
        # Planejamento
        (12, "Prioridade",              16,  COR_PLAN),
        (13, "Esforço (dias-aud)",      18,  COR_PLAN),
        (14, "Janela Planejada",        20,  COR_PLAN),
        (15, "Status",                  20,  COR_PLAN),
        (16, "Observações / Racional",  40,  COR_PLAN),
    ]
    _set_headers(ws, headers, row=3)
    ws.freeze_panes = "A4"

    end = 4 + N_ROWS - 1
    _data_rows(ws, N_COLS, 4, N_ROWS)

    _add_dv(ws, "C",  4, end, PROCESS_CODES)
    _add_dv(ws, "E",  4, end, TIPO_ENGAGEMENT)
    _add_dv(ws, "H",  4, end, IMPACTO_4)
    _add_dv(ws, "I",  4, end, PROB_4)
    _add_dv(ws, "K",  4, end, NIVEL_RISCO)
    _add_dv(ws, "L",  4, end, PRIORIDADE)
    _add_dv(ws, "O",  4, end, STATUS_ENGAGEMENT)

    _example_row(ws, [
        1, "Fechamento de Contrato", "FEC", "Checklist e baixa no sistema",
        "Assurance — Controles Internos", "2024", "Coordenador de Contratos",
        "Significativo (4) — >2% receita", "Provável (3) — mensal/trimestral",
        "Alto (12)", "Alto",
        "1 - Crítico", "5", "T1 / Jan-Mar",
        "planejamento", "",
    ], row=4)

    return ws


# ════════════════════════════════════════════════════════════════════════════
# ABA 2 — Plano Anual
# ════════════════════════════════════════════════════════════════════════════
def build_plano(wb):
    ws = wb.create_sheet("Plano Anual")
    ws.sheet_view.showGridLines = False

    N_STATIC = 7
    N_MESES  = 12
    N_EXTRA  = 2   # Status + Observações (Nível de Risco removido — duplica Universo)
    N_COLS   = N_STATIC + N_MESES + N_EXTRA  # 21

    _title_row(ws, "PLANO ANUAL DE AUDITORIA INTERNA  v3", N_COLS)
    _hint_row(ws,
        "Calendário: P=Planejado | E=Em Execução | C=Concluído | A=Adiado | X=Cancelado. "
        "Status segue states.yml: planejamento → execução → revisão_técnica → comunicação → monitoramento → arquivado.",
        N_COLS)

    # linha 3: grupo calendário
    cal_start = get_column_letter(N_STATIC + 1)
    cal_end   = get_column_letter(N_STATIC + N_MESES)
    ws.merge_cells(f"{cal_start}3:{cal_end}3")
    mg = ws.cell(row=3, column=N_STATIC + 1, value="CALENDÁRIO (marcar: P/E/C/A/X)")
    mg.fill = _fill(VERDE_HEADER); mg.font = _font(bold=True)
    mg.alignment = _center(); mg.border = _border_thin()
    for col in range(N_STATIC + 2, N_STATIC + N_MESES + 1):
        ws.cell(row=3, column=col).border = _border_thin()

    for col in range(1, N_STATIC + 1):
        ws.cell(row=3, column=col).fill = _fill(AZUL_HEADER)
        ws.cell(row=3, column=col).border = _border_thin()
    for col in range(N_STATIC + N_MESES + 1, N_COLS + 1):
        ws.cell(row=3, column=col).fill = _fill(AZUL_HEADER)
        ws.cell(row=3, column=col).border = _border_thin()
    ws.row_dimensions[3].height = 22

    # cabeçalhos linha 4
    static_headers = [
        (1,  "ID",                    8,  AZUL_HEADER),
        (2,  "Engagement",           35,  AZUL_HEADER),
        (3,  "Processo (code)",      16,  AZUL_HEADER),
        (4,  "Tipo",                 28,  AZUL_HEADER),
        (5,  "Prioridade",           14,  AZUL_HEADER),
        (6,  "Esforço (dias-aud)",   18,  AZUL_HEADER),
        (7,  "Responsável (Lead)",   22,  AZUL_HEADER),
    ]
    for col, label, width, fhex in static_headers:
        cell = ws.cell(row=4, column=col, value=label)
        cell.fill = _fill(fhex); cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True); cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(col)].width = width

    TRIMESTRE_FILLS = [
        ("DEEBF7", range(1, 4)),   # Q1 azul claro
        ("E2EFDA", range(4, 7)),   # Q2 verde claro
        ("FFF2CC", range(7, 10)),  # Q3 amarelo claro
        ("FCE4D6", range(10, 13)), # Q4 laranja claro
    ]
    for i, mes in enumerate(MESES, 1):
        col = N_STATIC + i
        fhex = next(h for h, rng in TRIMESTRE_FILLS if i in rng)
        cell = ws.cell(row=4, column=col, value=mes)
        cell.fill = _fill(fhex)
        cell.font = Font(bold=True, size=10, name="Arial", color="000000")
        cell.alignment = _center(); cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(col)].width = 8

    extra = [(N_COLS-1, "Status", 22), (N_COLS, "Observações", 35)]
    for col, label, width in extra:
        cell = ws.cell(row=4, column=col, value=label)
        cell.fill = _fill(AZUL_SUB); cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True); cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[4].height = 36
    ws.freeze_panes = "A5"

    end = 5 + N_ROWS - 1
    for row in range(5, end + 1):
        fill_row = _fill(CINZA_ALT) if row % 2 == 0 else _fill(BRANCO)
        for col in range(1, N_COLS + 1):
            cell = ws.cell(row=row, column=col)
            mes_i = col - N_STATIC
            if 1 <= mes_i <= 12:
                fhex = next(h for h, rng in TRIMESTRE_FILLS if mes_i in rng)
                cell.fill = _fill(fhex)
                cell.alignment = _center()
            else:
                cell.fill = fill_row
                cell.alignment = _left()
            cell.border = _border_thin()
            cell.font = Font(size=10, name="Arial", color="000000")

    _add_dv(ws, "C", 5, end, PROCESS_CODES)
    _add_dv(ws, "D", 5, end, TIPO_ENGAGEMENT)
    _add_dv(ws, "E", 5, end, PRIORIDADE)
    _add_dv(ws, get_column_letter(N_COLS - 1), 5, end, STATUS_ENGAGEMENT)

    return ws


# ════════════════════════════════════════════════════════════════════════════
# ABA 3 — Resumo Executivo
# ════════════════════════════════════════════════════════════════════════════
def build_resumo(wb):
    ws = wb.create_sheet("Resumo Executivo")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "RESUMO EXECUTIVO — PLANO ANUAL DE AUDITORIA  v3"
    t.fill  = _fill(AZUL_HEADER); t.font = _font(bold=True, size=13)
    t.alignment = _center(); t.border = _border_thin()
    ws.row_dimensions[1].height = 30

    widths_col = [35, 28, 35, 28, 35, 28]
    for i, w in enumerate(widths_col, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    sections = [
        ("ESTATÍSTICAS DO PLANO", AZUL_SUB, [
            ("Total de engagements planejados",   ""),
            ("Total de dias-auditoria alocados",  ""),
            ("Engagements concluídos",            ""),
            ("Engagements em execução",           ""),
            ("Engagements adiados / cancelados",  ""),
        ]),
        ("DISTRIBUIÇÃO POR PRIORIDADE", VERDE_HEADER, [
            ("1 - Crítico",    ""),
            ("2 - Alto",       ""),
            ("3 - Moderado",   ""),
            ("4 - Baixo",      ""),
        ]),
        ("DISTRIBUIÇÃO POR NÍVEL DE RISCO", ROXO_HEADER, [
            ("Crítico (score 13–16)",  ""),
            ("Alto (score 10–12)",     ""),
            ("Moderado (score 4–9)",   ""),
            ("Baixo (score 1–3)",      ""),
        ]),
        ("CAPACIDADE DA EQUIPE", "2E75B6", [
            ("Auditores disponíveis (FTE)",  ""),
            ("Dias úteis no ano",            ""),
            ("Capacidade total (dias-aud)",  ""),
            ("Capacidade utilizada (%)",     ""),
            ("Reserva de capacidade",        ""),
        ]),
        ("COBERTURA DO UNIVERSO", "7B2C2C", [
            ("Total de processos no universo",       ""),
            ("Processos cobertos pelo plano",        ""),
            ("% de cobertura",                       ""),
            ("Processos sem cobertura — risco Alto+",""),
        ]),
    ]

    r = 3
    for section_title, fill_hex, campos in sections:
        ws.merge_cells(f"A{r}:F{r}")
        hdr = ws.cell(row=r, column=1, value=section_title)
        hdr.fill = _fill(fill_hex); hdr.font = _font(bold=True, size=11)
        hdr.alignment = _left(wrap=False); hdr.border = _border_thin()
        ws.row_dimensions[r].height = 24; r += 1

        for campo, default_val in campos:
            cl = ws.cell(row=r, column=1, value=campo)
            cl.fill = _fill("EBF3FB"); cl.font = _font(bold=True, color="000000", size=10)
            cl.border = _border_thin(); cl.alignment = _left(wrap=False)

            ws.merge_cells(f"B{r}:C{r}")
            vc = ws.cell(row=r, column=2, value=default_val)
            vc.fill = _fill(BRANCO); vc.border = _border_thin()
            vc.alignment = _center()
            vc.font = Font(size=11, bold=True, name="Arial", color="000000")

            for col in range(4, 7):
                ws.cell(row=r, column=col).fill = _fill(CINZA_ALT)
                ws.cell(row=r, column=col).border = _border_thin()
            ws.row_dimensions[r].height = 20; r += 1
        r += 1

    return ws


# ════════════════════════════════════════════════════════════════════════════
# ABA 4 — _Ref_Critérios
# ════════════════════════════════════════════════════════════════════════════
def build_ref_criterios(wb):
    ws = wb.create_sheet("_Ref_Critérios")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "REFERÊNCIA — CRITÉRIOS DE PRIORIZAÇÃO + HEAT MAP + ESTADOS"
    t.fill  = _fill(AZUL_HEADER); t.font = _font(bold=True, size=11)
    t.alignment = _center(); t.border = _border_thin()
    ws.row_dimensions[1].height = 28

    ws["A2"] = ("Fonte: _method-wiki/concepts/risk-scoring-foundations.md — COSO ERM 2017 | "
                "states.yml | IIA IPPF 2024")
    ws["A2"].font = Font(size=9, name="Arial", color="595959", italic=True)

    r = 4

    # ── Tabela 1: Critérios de priorização ───────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    h = ws.cell(row=r, column=1, value="CRITÉRIOS DE PRIORIZAÇÃO DO UNIVERSO AUDITÁVEL (pesos sugeridos)")
    h.fill = _fill(AZUL_SUB); h.font = _font(bold=True); h.alignment = _center()
    h.border = _border_thin(); ws.row_dimensions[r].height = 22; r += 1

    crit_headers = ["Critério", "Peso (%)", "Como Avaliar", "Escala", "Fonte"]
    crit_widths  = [28, 10, 52, 30, 30]
    for ci, (hdr, w) in enumerate(zip(crit_headers, crit_widths), 1):
        cell = ws.cell(row=r, column=ci, value=hdr)
        cell.fill = _fill(AZUL_SUB); cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True); cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[r].height = 28; r += 1

    criterios = [
        ("Impacto Potencial (magnitude)", "35%",
         "Magnitude do dano financeiro, reputacional ou regulatório caso o risco se materialize.",
         "1=Insignificante ... 4=Significativo",
         "COSO ERM 2017 / referencial metodológico Cap.3"),
        ("Probabilidade de Ocorrência", "30%",
         "Histórico de incidentes, qualidade do controle atual, complexidade operacional.",
         "1=Improvável ... 4=Quase Certo",
         "COSO ERM 2017 / referencial metodológico Cap.3"),
        ("Qualidade do Ambiente de Controle", "20%",
         "Avaliação do controle existente: desenhado, operando, evidenciado.",
         "1=Robusto ... 4=Inexistente",
         "COSO P10-P12"),
        ("Tempo desde Última Auditoria", "10%",
         "Quanto tempo sem cobertura de auditoria interna.",
         "1=<1 ano ... 4=>4 anos",
         "IIA IPPF §2520"),
        ("Pressão Regulatória / Conformidade", "5%",
         "Requisito normativo explícito, auditoria externa pendente ou solicitação do board.",
         "0=Não aplicável / 1=Aplicável",
         "IIA IPPF / SOX / LGPD"),
    ]
    for row_data in criterios:
        fill = _fill(CINZA_ALT) if r % 2 == 0 else _fill(BRANCO)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.fill = fill; cell.border = _border_thin()
            cell.alignment = _left(wrap=(ci in (3, 4, 5)))
            cell.font = Font(size=10, name="Arial", color="000000", bold=(ci == 1))
        ws.row_dimensions[r].height = 40; r += 1

    r += 1

    # ── Tabela 2: Heat map 4×4 ────────────────────────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    h2 = ws.cell(row=r, column=1, value="HEAT MAP 4×4 — SCORE DE RISCO INERENTE (referência rápida)")
    h2.fill = _fill(AZUL_SUB); h2.font = _font(bold=True); h2.alignment = _center()
    h2.border = _border_thin(); ws.row_dimensions[r].height = 22; r += 1

    ws.cell(row=r, column=1, value="Impacto \\ Prob.").fill = _fill(CINZA_TITULO)
    ws.cell(row=r, column=1).font = Font(bold=True, size=9, name="Arial", color="000000")
    ws.cell(row=r, column=1).alignment = _center(wrap=True)
    ws.cell(row=r, column=1).border = _border_thin()

    prob_short = ["Improvável\n(1)", "Possível\n(2)", "Provável\n(3)", "Quase Certo\n(4)"]
    for j, lp in enumerate(prob_short, 2):
        c = ws.cell(row=r, column=j, value=lp)
        c.fill = _fill(AZUL_SUB); c.font = _font(bold=True, size=9)
        c.alignment = _center(wrap=True); c.border = _border_thin()
    ws.row_dimensions[r].height = 30; r += 1

    SCORE_COLORS = {
        "Baixo":    ("92D050", "000000"),
        "Moderado": ("FFFF00", "000000"),
        "Alto":     ("FF9900", "000000"),
        "Crítico":  ("FF0000", BRANCO),
    }
    imp_short = ["Insignif. (1)", "Baixo (2)", "Moderado (3)", "Signific. (4)"]
    for i, li in enumerate(imp_short, 1):
        c0 = ws.cell(row=r, column=1, value=li)
        c0.fill = _fill(CINZA_TITULO)
        c0.font = Font(bold=True, size=9, name="Arial", color="000000")
        c0.alignment = _left(wrap=False); c0.border = _border_thin()
        ws.row_dimensions[r].height = 22
        for j in range(1, 5):
            label, score = SCORE_BANDS[(i, j)]
            fhex, fc = SCORE_COLORS[label]
            cell = ws.cell(row=r, column=j + 1, value=f"{label} ({score})")
            cell.fill = _fill(fhex)
            cell.font = Font(bold=True, size=9, name="Arial", color=fc)
            cell.alignment = _center(); cell.border = _border_thin()
        r += 1

    ws.merge_cells(f"A{r}:F{r}")
    ws.cell(row=r, column=1,
            value="Bandas: 1–3=Baixo | 4–9=Moderado | 10–12=Alto | 13–16=Crítico").font = \
        Font(size=9, name="Arial", color="595959", italic=True, bold=True)
    ws.row_dimensions[r].height = 18; r += 2

    # ── Tabela 3: Estados canônicos do engagement ─────────────────────────
    ws.merge_cells(f"A{r}:F{r}")
    h3 = ws.cell(row=r, column=1, value="ESTADOS CANÔNICOS DO ENGAGEMENT (states.yml)")
    h3.fill = _fill(VERDE_HEADER); h3.font = _font(bold=True); h3.alignment = _center()
    h3.border = _border_thin(); ws.row_dimensions[r].height = 22; r += 1

    est_headers = ["Estado", "Significado Operacional"]
    for ci, hdr in enumerate(est_headers, 1):
        cell = ws.cell(row=r, column=ci, value=hdr)
        ws.merge_cells(f"B{r}:F{r}") if ci == 2 else None
        cell.fill = _fill(VERDE_HEADER); cell.font = _font(bold=True)
        cell.alignment = _center(wrap=True); cell.border = _border_thin()
    ws.row_dimensions[r].height = 24; r += 1

    estados = [
        ("planejamento",    "Escopo, riscos e programa de trabalho definidos; fieldwork não iniciado."),
        ("execução",        "Fieldwork em andamento; testes e coleta de evidências ativos."),
        ("revisão_técnica", "Auditor sênior / gerente revisando workpapers e conclusões."),
        ("comunicação",     "Achados comunicados à gestão; draft ou relatório final emitido."),
        ("monitoramento",   "Acompanhando implementação dos planos de ação."),
        ("arquivado",       "Engagement encerrado; planos de ação aceitos ou verificados."),
        ("adiado",          "Engagement reprogramado; razão documentada."),
        ("cancelado",       "Engagement removido do plano; razão documentada e aprovada."),
        ("não_iniciado",    "Planejado mas ainda sem ação. Default para novos registros."),
    ]
    for estado, significado in estados:
        fill = _fill(CINZA_ALT) if r % 2 == 0 else _fill(BRANCO)
        ws.cell(row=r, column=1, value=estado).fill = fill
        ws.cell(row=r, column=1).font = Font(size=10, name="Arial", color="000000", bold=True)
        ws.cell(row=r, column=1).border = _border_thin()
        ws.cell(row=r, column=1).alignment = _left()
        ws.merge_cells(f"B{r}:F{r}")
        ws.cell(row=r, column=2, value=significado).fill = fill
        ws.cell(row=r, column=2).font = Font(size=10, name="Arial", color="000000")
        ws.cell(row=r, column=2).border = _border_thin()
        ws.cell(row=r, column=2).alignment = _left(wrap=True)
        ws.row_dimensions[r].height = 22; r += 1

    return ws


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════
def main():
    script_dir  = Path(__file__).parent
    repo_root   = script_dir.parent
    default_out = repo_root / "templates" / "planilhas" / "Plano_Anual_Auditoria_v3_template.xlsx"
    out_path = Path(sys.argv[1]) if len(sys.argv) > 1 else default_out

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_universo(wb)
    build_plano(wb)
    build_resumo(wb)
    build_ref_criterios(wb)

    wb.save(out_path)
    print(f"Template gerado: {out_path}")


if __name__ == "__main__":
    main()
